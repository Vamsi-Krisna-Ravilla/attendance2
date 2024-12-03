import streamlit as st
import pandas as pd
from datetime import datetime
import hashlib
import openpyxl
from openpyxl.styles import Alignment
import io
import numpy as np


# Configure Streamlit page - must be the first Streamlit command
st.set_page_config(page_title="Attendance Management System", layout="wide")

# Admin credentials
ADMIN_CREDENTIALS = {
    "a": hashlib.sha256("a".encode()).hexdigest()
}

def check_login(username, password, is_admin=False):
    """Verify login credentials"""
    try:
        if is_admin:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            return ADMIN_CREDENTIALS.get(username) == hashed_password
        else:
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            return any((df_faculty['Faculty Name'] == username) & 
                      (df_faculty['Password'] == password))
    except Exception as e:
        st.error(f"Login error: {str(e)}")
        return False

def get_section_subjects(section):
    """Get subjects mapped to a specific section"""
    try:
        df_mapping = pd.read_excel('attendance.xlsx', sheet_name='Section-Subject-Mapping')
        section_row = df_mapping[df_mapping['Section'] == section].iloc[0]
        subjects = [s.strip() for s in str(section_row['Subject Names']).split('\n') if s.strip()]
        return subjects
    except Exception as e:
        st.error(f"Error getting subjects: {str(e)}")
        return []

def get_student_data(section):
    """Get student data for a specific section"""
    try:
        # Read from sheet with (O) prefix to get original data
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        # Only return HT Number and Student Name columns
        return df[['HT Number', 'Student Name']].fillna('')
    except Exception as e:
        st.error(f"Error getting student data: {str(e)}")
        return None

# Add new function for faculty workload
def get_faculty_workload(username):
    """Calculate faculty workload"""
    try:
        all_sheets = pd.ExcelFile('attendance.xlsx').sheet_names
        section_sheets = [s for s in all_sheets if s.startswith('(O)')]
        
        total_periods = 0
        workload_data = []
        
        for sheet in section_sheets:
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if period in df.columns:
                    for _, row in df.iterrows():
                        if pd.notna(row[period]):
                            entries = str(row[period]).split('\n')
                            for entry in entries:
                                if username in entry:  # Check if faculty took this class
                                    date, time, status, faculty, subject = entry.split('_')
                                    workload_data.append({
                                        'Date': date,
                                        'Time': time,
                                        'Period': period,
                                        'Section': sheet.replace('(O)', ''),
                                        'Subject': subject
                                    })
                                    total_periods += 1
        
        return total_periods, pd.DataFrame(workload_data)
    except Exception as e:
        st.error(f"Error calculating workload: {str(e)}")
        return 0, pd.DataFrame()
def check_duplicate_attendance(section, period, date_str):
    """Check if attendance is already marked for given section, period and date"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        if period in df.columns:
            for value in df[period].dropna():
                for entry in str(value).split('\n'):
                    if entry.startswith(date_str):
                        return True
        return False
    except Exception as e:
        st.error(f"Error checking duplicate attendance: {str(e)}")
        return True

def mark_attendance(section, period, attendance_data, faculty_name, subject):
    """Mark attendance with proper Excel formatting"""
    try:
        date_str = datetime.now().strftime('%d/%m/%Y')
        time_str = datetime.now().strftime('%I:%M%p')
        
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        
        with pd.ExcelWriter('attendance.xlsx', 
                           mode='a', 
                           if_sheet_exists='overlay', 
                           engine='openpyxl') as writer:
            # Write data
            for ht_number, status in attendance_data.items():
                attendance_value = f"{date_str}_{time_str}_{status}_{faculty_name}_{subject}"
                current_value = df.loc[df['HT Number'] == ht_number, period].iloc[0]
                df.loc[df['HT Number'] == ht_number, period] = (
                    f"{current_value}\n{attendance_value}" if pd.notna(current_value) and current_value 
                    else attendance_value
                )
            
            # Write to Excel
            df.to_excel(writer, sheet_name=f'(O){section}', index=False)
            
            # Get the worksheet
            worksheet = writer.sheets[f'(O){section}']
            
            # Format cells for multiline
            for row in worksheet.iter_rows():
                for cell in row:
                    # Set text wrap and auto-adjust row height
                    cell.alignment = openpyxl.styles.Alignment(
                        wrap_text=True,
                        vertical='top'
                    )
                    # Adjust row height
                    worksheet.row_dimensions[cell.row].height = None  # Auto-height
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value).split('\n')[0])  # Check first line length
                        )
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        return True
    except Exception as e:
        st.error(f"Error marking attendance: {str(e)}")
        return False


def get_attendance_stats(section, from_date=None, to_date=None):
    """Calculate attendance statistics with subject-wise breakdown"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        subjects = get_section_subjects(section)
        
        stats = []
        for _, row in df.iterrows():
            student_stats = {
                'HT Number': row['HT Number'],
                'Student Name': row['Student Name']
            }
            
            # Initialize subject-wise counters
            subject_stats = {subject: {'present': 0, 'total': 0} for subject in subjects}
            
            # Process attendance data
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if pd.notna(row[period]) and row[period]:
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        try:
                            date_str, _, status, _, subject = entry.split('_')
                            entry_date = datetime.strptime(date_str, '%d/%m/%Y')
                            
                            if ((not from_date or entry_date >= datetime.strptime(from_date, '%Y-%m-%d')) and
                                (not to_date or entry_date <= datetime.strptime(to_date, '%Y-%m-%d'))):
                                if subject in subject_stats:
                                    subject_stats[subject]['total'] += 1
                                    if status == 'P':
                                        subject_stats[subject]['present'] += 1
                        except:
                            continue
            
            # Calculate percentages
            total_present = 0
            total_classes = 0
            
            for subject in subjects:
                present = subject_stats[subject]['present']
                total = subject_stats[subject]['total']
                total_present += present
                total_classes += total
                
                if total > 0:
                    percentage = (present / total) * 100
                else:
                    percentage = 0
                
                student_stats.update({
                    f"{subject} Present": present,
                    f"{subject} Total": total,
                    f"{subject} %": round(percentage, 2)
                })
            
            # Calculate overall percentage
            if total_classes > 0:
                overall_percentage = (total_present / total_classes) * 100
            else:
                overall_percentage = 0
                
            student_stats.update({
                'Total Present': total_present,
                'Total Classes': total_classes,
                'Overall %': round(overall_percentage, 2)
            })
            
            stats.append(student_stats)
            
        return pd.DataFrame(stats)
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None

def faculty_page():
    st.title(f"Welcome, {st.session_state.username}")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio("Select", ["Mark Attendance", "View Statistics", "Student Reports", "Subject Analysis", "My Workload"])
        
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    if page == "My Workload":
        st.subheader("Faculty Workload Summary")
        
        # Calculate workload
        total_periods, workload_df = get_faculty_workload(st.session_state.username)
        
        # Display metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Classes Taken", total_periods)
        with col2:
            total_hours = total_periods  # Each period is 1 hour
            st.metric("Total Hours", f"{total_hours:,}")
        with col3:
            if not workload_df.empty:
                unique_days = workload_df['Date'].nunique()
                st.metric("Days Engaged", unique_days)
        
        # Show detailed workload data
        if not workload_df.empty:
            st.subheader("Detailed Class Records")
            # Sort by date and time
            workload_df = workload_df.sort_values(['Date', 'Time'], ascending=[False, False])
            st.dataframe(workload_df)
            
            if st.button("Download Workload Report"):
                csv = workload_df.to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name=f"faculty_workload_{st.session_state.username}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
            
            # Show subject-wise breakdown
            st.subheader("Subject-wise Classes")
            subject_counts = workload_df['Subject'].value_counts()
            subject_df = pd.DataFrame({
                'Subject': subject_counts.index,
                'Classes Taken': subject_counts.values
            })
            st.dataframe(subject_df)
        else:
            st.info("No classes recorded yet")
            
    elif page == "Mark Attendance":
        st.subheader("Mark Attendance")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            sections = [s.replace('(O)', '') for s in pd.ExcelFile('attendance.xlsx').sheet_names 
                       if s.startswith('(O)')]
            section = st.selectbox("Select Section", sections)
        
        with col2:
            subjects = get_section_subjects(section) if section else []
            subject = st.selectbox("Select Subject", subjects)
        
        with col3:
            period = st.selectbox("Select Period", ['P1', 'P2', 'P3', 'P4', 'P5', 'P6'])
        
        if section and subject and period:
            df_students = get_student_data(section)
            if df_students is not None:
                attendance_data = {}
                
                # Select all/none buttons
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("Select All"):
                        st.session_state.select_all = True
                with col2:
                    if st.button("Select None"):
                        st.session_state.select_all = False
                
                st.write("### Students")
                for _, student in df_students.iterrows():
                    col1, col2, col3 = st.columns([2, 2, 1])
                    with col1:
                        st.write(student['HT Number'])
                    with col2:
                        st.write(student['Student Name'])
                    with col3:
                        default_value = getattr(st.session_state, 'select_all', True)
                        status = 'P' if st.checkbox("Present", key=student['HT Number'], value=default_value) else 'A'
                        attendance_data[student['HT Number']] = status
                
                if st.button("Submit Attendance"):
                    if mark_attendance(section, period, attendance_data, st.session_state.username, subject):
                        st.success("Attendance marked successfully!")
                    else:
                        st.error("Failed to mark attendance")
    
    elif page == "View Statistics":
        st.subheader("Attendance Statistics")
        
        sections = [s.replace('(O)', '') for s in pd.ExcelFile('attendance.xlsx').sheet_names 
                   if s.startswith('(O)')]
        section = st.selectbox("Select Section", sections)
        
        col1, col2 = st.columns(2)
        with col1:
            from_date = st.date_input("From Date")
        with col2:
            to_date = st.date_input("To Date")
        
        if section:
            stats_df = get_attendance_stats(section, 
                                         from_date.strftime('%Y-%m-%d') if from_date else None,
                                         to_date.strftime('%Y-%m-%d') if to_date else None)
            
            if stats_df is not None:
                st.write("### Overall Statistics")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.metric("Total Students", len(stats_df))
                with col2:
                    avg_attendance = stats_df['Overall %'].mean()
                    st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                with col3:
                    below_75 = len(stats_df[stats_df['Overall %'] < 75])
                    st.metric("Students Below 75%", below_75)
                
                st.write("### Student-wise Statistics")
                st.dataframe(stats_df)
                
                if st.button("Download Report"):
                    csv = stats_df.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"attendance_stats_{section}_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv"
                    )
    
    elif page == "Student Reports":
        st.subheader("Individual Student Reports")
        
        sections = [s.replace('(O)', '') for s in pd.ExcelFile('attendance.xlsx').sheet_names 
                   if s.startswith('(O)')]
        section = st.selectbox("Select Section", sections)
        
        if section:
            df_students = get_student_data(section)
            if df_students is not None:
                student = st.selectbox("Select Student", 
                                     df_students['HT Number'].tolist(),
                                     format_func=lambda x: f"{x} - {df_students[df_students['HT Number']==x]['Student Name'].iloc[0]}")
                
                if student:
                    st.write(f"### Attendance Report for {student}")
                    student_data = df_students[df_students['HT Number'] == student].iloc[0]
                    
                    attendance_data = []
                    for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                        if pd.notna(student_data[period]) and student_data[period]:
                            entries = student_data[period].split('\n')
                            for entry in entries:
                                if entry.strip():
                                    try:
                                        date, time, status, faculty, subject = entry.split('_')
                                        attendance_data.append({
                                            'Date': date,
                                            'Time': time,
                                            'Period': period,
                                            'Status': status,
                                            'Faculty': faculty,
                                            'Subject': subject
                                        })
                                    except:
                                        continue
                    
                    if attendance_data:
                        df_attendance = pd.DataFrame(attendance_data)
                        df_attendance = df_attendance.sort_values('Date', ascending=False)
                        st.dataframe(df_attendance)
                        
                        if st.button("Download Student Report"):
                            csv = df_attendance.to_csv(index=False)
                            st.download_button(
                                label="Download CSV",
                                data=csv,
                                file_name=f"student_report_{student}_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv"
                            )
                    else:
                        st.info("No attendance records found")
    
    else:  # Subject Analysis
        st.subheader("Subject-wise Analysis")
        
        sections = [s.replace('(O)', '') for s in pd.ExcelFile('attendance.xlsx').sheet_names 
                   if s.startswith('(O)')]
        section = st.selectbox("Select Section", sections)
        
        if section:
            subjects = get_section_subjects(section)
            subject = st.selectbox("Select Subject", subjects)
            
            if subject:
                df_students = get_student_data(section)
                if df_students is not None:
                    subject_stats = []
                    
                    for _, student in df_students.iterrows():
                        present = 0
                        total = 0
                        
                        for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                            if pd.notna(student[period]) and student[period]:
                                entries = str(student[period]).split('\n')
                                for entry in entries:
                                    if entry.strip() and subject in entry:
                                        total += 1
                                        if '_P_' in entry:
                                            present += 1
                        
                        if total > 0:
                            percentage = (present / total) * 100
                        else:
                            percentage = 0
                            
                        subject_stats.append({
                            'HT Number': student['HT Number'],
                            'Student Name': student['Student Name'],
                            'Classes Attended': present,
                            'Total Classes': total,
                            'Attendance %': round(percentage, 2)
                        })
                    
                    df_stats = pd.DataFrame(subject_stats)
                    
                    st.write("### Subject Statistics")
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        st.metric("Average Attendance", f"{df_stats['Attendance %'].mean():.2f}%")
                    with col2:
                        st.metric("Total Classes", df_stats['Total Classes'].max())
                    with col3:
                        below_75 = len(df_stats[df_stats['Attendance %'] < 75])
                        st.metric("Students Below 75%", below_75)
                    
                    st.write("### Student-wise Subject Statistics")
                    st.dataframe(df_stats)
                    
                    if st.button("Download Subject Report"):
                        csv = df_stats.to_csv(index=False)
                        st.download_button(
                            label="Download CSV",
                            data=csv,
                            file_name=f"subject_stats_{section}_{subject}_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )

def get_column_width(col_name, values):
    try:
        max_length = max(
            max(len(str(val)) for val in values if val is not None),
            len(str(col_name))
        )
        return min(max_length * 10, 300)
    except:
        return 150

def admin_page():
    st.title("Admin Dashboard")
    
    with st.sidebar:
        st.header("Data Management")
        sheet = st.selectbox("Select Sheet", pd.ExcelFile('attendance.xlsx').sheet_names)
        
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    st.subheader(f"Edit {sheet}")
    
    try:
        # Read data
        df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
        
        # Convert all columns to string and handle special formatting
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].astype(float).astype(str)
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            else:
                df[col] = df[col].astype(str)
        
        df = df.fillna('')
        
        # Remove empty rows
        df = df.loc[df.replace('', np.nan).notna().any(axis=1)].copy()
        df = df.reset_index(drop=True)
        
        # Calculate column heights based on content
        max_lines = {col: 1 for col in df.columns}
        for col in df.columns:
            for val in df[col]:
                if pd.notna(val) and val:
                    lines = str(val).count('\n') + 1
                    max_lines[col] = max(max_lines[col], lines)
    
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return

    # Configure columns with appropriate heights
    column_config = {}
    for col in df.columns:
        # Base height on content
        height = max(100, min(400, max_lines[col] * 35))
        
        # Different widths for different column types
        if col in ['Faculty Name', 'Credential', 'Student Name', 'HT Number']:
            width = 150
        else:
            width = 300
            
        column_config[str(col)] = st.column_config.TextColumn(
            str(col),
            width=width,
            help=f"Enter {col}",
            max_chars=None
        )

    # Add filters
    with st.expander("Show Filters", expanded=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            search_text = st.text_input("Search", placeholder="Search in any column...")
            if search_text:
                mask = df.apply(lambda x: x.str.contains(search_text, case=False, na=False)).any(axis=1)
                df = df[mask]
        
        with col2:
            filter_column = st.selectbox("Filter by column", ["None"] + list(df.columns))
            if filter_column != "None":
                unique_values = sorted(df[filter_column].unique().tolist())
                selected_values = st.multiselect(
                    f"Select {filter_column} values",
                    unique_values,
                    default=unique_values
                )
                if selected_values:
                    df = df[df[filter_column].isin(selected_values)]
        
        with col3:
            sort_column = st.selectbox("Sort by", ["None"] + list(df.columns))
            if sort_column != "None":
                sort_order = st.radio("Sort order", ["Ascending", "Descending"])
                df = df.sort_values(by=sort_column, ascending=(sort_order == "Ascending"))

    # Add new row button
    if st.button("Add New Row"):
        new_row = pd.DataFrame([['' for _ in df.columns]], columns=df.columns)
        df = pd.concat([df, new_row], ignore_index=True)

    # Display editor with multiline support
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        column_config=column_config,
        hide_index=True,
        height=min(800, len(df) * 50 + 100),  # Adjust height based on content
        key=f"editor_{sheet}"
    )

    # Remove empty rows before saving
    edited_df = edited_df.loc[edited_df.replace('', np.nan).notna().any(axis=1)].copy()

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Save Changes"):
            try:
                import os
                file_path = 'attendance.xlsx'
                
                # Check if file is writable
                if not os.access(file_path, os.W_OK):
                    st.error("Cannot write to file. Please check file permissions.")
                    return
                
                # Check if file is locked
                try:
                    with open(file_path, 'a') as f:
                        pass
                except IOError:
                    st.error("File is locked. Please close it in other programs and try again.")
                    return
                
                with pd.ExcelWriter(
                    file_path, 
                    mode='a', 
                    engine='openpyxl',
                    if_sheet_exists='overlay'
                ) as writer:
                    edited_df.to_excel(writer, sheet_name=sheet, index=False)
                    
                    worksheet = writer.sheets[sheet]
                    
                    # Format cells for multiline
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=True,
                                vertical='top'
                            )
                            
                            # Set row height based on content
                            if cell.value:
                                lines = str(cell.value).count('\n') + 1
                                current_height = worksheet.row_dimensions[cell.row].height or 15
                                worksheet.row_dimensions[cell.row].height = max(current_height, lines * 15)
                    
                    # Set column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, max(len(line) for line in str(cell.value).split('\n')))
                        adjusted_width = min(50, max(12, max_length + 2))
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                st.success("Changes saved successfully!")
                st.rerun()
            except PermissionError:
                st.error("Cannot save changes. Please check if the file is open in another program.")
            except Exception as e:
                st.error(f"Error saving changes: {str(e)}")
                st.info("Try closing the Excel file if it's open in another program.")


def main():
    if 'logged_in' not in st.session_state:
        st.title("Login")
        
        # Show sample credentials
        with st.expander("View Demo Credentials"):
            st.info("""
            Admin Login:
            - Username: admin
            - Password: admin123
            
            Faculty Login:
            - Username: faculty1
            - Password: pass123
            """)
        
        login_type = st.radio("Select Login Type", ["Faculty", "Admin"])
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        
        if st.button("Login"):
            if check_login(username, password, login_type == "Admin"):
                st.session_state.logged_in = True
                st.session_state.is_admin = (login_type == "Admin")
                st.session_state.username = username
                st.rerun()
            else:
                st.error("Invalid credentials")
    else:
        if st.session_state.is_admin:
            admin_page()
        else:
            faculty_page()

if __name__ == "__main__":
    main()