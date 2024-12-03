import streamlit as st
import pandas as pd
from datetime import datetime
import hashlib
import openpyxl
from openpyxl.styles import Alignment
import io
import numpy as np


# Configure Streamlit page - must be the first Streamlit command
st.set_page_config(page_title="Attendance Management System", layout="wide")

# Admin credentials
ADMIN_CREDENTIALS = {
    "a": hashlib.sha256("a".encode()).hexdigest()
}

def check_login(username, password, is_admin=False):
    """Verify login credentials"""
    try:
        if is_admin:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            return ADMIN_CREDENTIALS.get(username) == hashed_password
        else:
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            return any((df_faculty['Faculty Name'] == username) & 
                      (df_faculty['Password'] == password))
    except Exception as e:
        st.error(f"Login error: {str(e)}")
        return False

def get_section_subjects(section):
    """Get subjects mapped to a specific section"""
    try:
        df_mapping = pd.read_excel('attendance.xlsx', sheet_name='Section-Subject-Mapping')
        section_row = df_mapping[df_mapping['Section'] == section].iloc[0]
        subjects = [s.strip() for s in str(section_row['Subject Names']).split('\n') if s.strip()]
        return subjects
    except Exception as e:
        st.error(f"Error getting subjects: {str(e)}")
        return []

def get_sections():
    """Get manipulated sections (without O prefix) for attendance marking"""
    try:
        all_sheets = pd.ExcelFile('attendance.xlsx').sheet_names
        return [s for s in all_sheets if not s.startswith('(O)')]
    except Exception as e:
        st.error(f"Error getting sections: {str(e)}")
        return []

def mark_attendance(section, period, attendance_data, faculty_name, subject):
    """Mark attendance in original sections"""
    try:
        date_str = datetime.now().strftime('%d/%m/%Y')
        time_str = datetime.now().strftime('%I:%M%p')
        
        # Group students by their original sections
        original_sections = {}
        for ht_number, data in attendance_data.items():
            # Remove the "Original: " prefix and use just the section name
            orig_section = data['original_section'].replace("Original: ", "")
            if orig_section not in original_sections:
                original_sections[orig_section] = {}
            original_sections[orig_section][ht_number] = data['status']
        
        # Mark attendance in each original section
        for orig_section, students in original_sections.items():
            # Read the original section sheet
            df = pd.read_excel('attendance.xlsx', sheet_name=orig_section)
            
            with pd.ExcelWriter('attendance.xlsx', 
                           mode='a', 
                           if_sheet_exists='overlay', 
                           engine='openpyxl') as writer:
                for ht_number, status in students.items():
                    # Format attendance entry like existing entries
                    attendance_value = f"{date_str}_{time_str}_{status}_{faculty_name}_{subject}"
                    row_mask = df['HT Number'] == ht_number
                    if row_mask.any():
                        current_value = df.loc[row_mask, period].iloc[0]
                        df.loc[row_mask, period] = (
                            f"{current_value}\n{attendance_value}" if pd.notna(current_value) and current_value 
                            else attendance_value
                        )
                
                # Save to original section
                df.to_excel(writer, sheet_name=orig_section, index=False)
                
                # Format worksheet
                worksheet = writer.sheets[orig_section]
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(
                            wrap_text=True,
                            vertical='top'
                        )
                    worksheet.row_dimensions[row[0].row].height = None  # Auto-adjust height
                
                # Set column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            max_length = max(max_length, max(len(line) for line in lines))
                    adjusted_width = min(50, max(12, max_length + 2))
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return True
    except Exception as e:
        st.error(f"Error marking attendance: {str(e)}")
        return False

def get_student_data(section):
    """Get student data for a manipulated section"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=section)
        students_df = df[['HT Number', 'Student Name', 'Original Section']].fillna('')
        # Format original section display without duplication
        students_df['Original Section'] = students_df['Original Section'].apply(
            lambda x: f"Original: {x}" if not x.startswith("Original:") else x
        )
        return students_df
    except Exception as e:
        st.error(f"Error getting student data: {str(e)}")
        return None

def faculty_page():
    st.title(f"Welcome, {st.session_state.username}")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio(
            "Select", 
            ["Mark Attendance", "View Statistics", "Student Reports", 
             "Subject Analysis", "My Workload"]
        )

        if page == "Mark Attendance":
            # Period selection
            st.session_state.period = st.selectbox(
                "Select Period",
                options=[''] + ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']
            )
            
            # Section selection - any manipulated section
            sections = get_sections()
            st.session_state.sections = st.multiselect(
                "Select Section(s)",
                options=sections,
                max_selections=2,
                help="Select up to 2 sections"
            )
            
            # Show available subjects for selected section
            if st.session_state.sections:
                all_subjects = []
                for section in st.session_state.sections:
                    subjects = get_section_subjects(section)
                    all_subjects.extend(subjects)
                # Remove duplicates and sort
                unique_subjects = sorted(list(set(all_subjects)))
                
                st.session_state.subject = st.selectbox(
                    "Select Subject",
                    options=unique_subjects if unique_subjects else [''],
                    help="Select the subject being taught"
                )
        
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    if page == "Mark Attendance":
        if not (st.session_state.get('period') and st.session_state.get('sections') 
                and st.session_state.get('subject')):
            st.info("Please select Period, Section(s), and Subject from the sidebar")
            return
            
        mark_attendance_page()
        
    elif page == "View Statistics":
        view_statistics_page()
        
    elif page == "Student Reports":
        student_reports_page()
        
    elif page == "Subject Analysis":
        subject_analysis_page()
        
    else:  # Workload Analysis
        workload_analysis_page()


def mark_attendance_page():
    """Page for marking attendance"""
    section = st.session_state.sections[0] if st.session_state.sections else None
    subject = st.session_state.subject
    period = st.session_state.period

    if section and subject and period:
        df_students = get_student_data(section)
        if df_students is not None:
            attendance_data = {}
            
            # Select all/none buttons
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Select All"):
                    st.session_state.select_all = True
            with col2:
                if st.button("Select None"):
                    st.session_state.select_all = False
            
            st.write("### Students")
            for _, student in df_students.iterrows():
                col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                with col1:
                    st.write(student['HT Number'])
                with col2:
                    st.write(student['Student Name'])
                with col3:
                    st.write(student['Original Section'])
                with col4:
                    default_value = getattr(st.session_state, 'select_all', True)
                    status = 'P' if st.checkbox("Present", key=student['HT Number'], value=default_value) else 'A'
                    attendance_data[student['HT Number']] = {
                        'status': status,
                        'original_section': student['Original Section']
                    }
            
            if st.button("Submit Attendance"):
                if mark_attendance(section, period, attendance_data, st.session_state.username, subject):
                    st.success("Attendance marked successfully!")
                else:
                    st.error("Failed to mark attendance")


def student_reports_page():
    """Page for individual student reports"""
    st.subheader("Individual Student Reports")
    
    sections = get_sections()
    section = st.selectbox("Select Section", sections)
    
    if section:
        df_students = get_student_data(section)
        if df_students is not None:
            student = st.selectbox(
                "Select Student", 
                df_students['HT Number'].tolist(),
                format_func=lambda x: f"{x} - {df_students[df_students['HT Number']==x]['Student Name'].iloc[0]}"
            )
            
            if student:
                student_data = df_students[df_students['HT Number'] == student].iloc[0]
                
                st.write(f"### Attendance Report for {student}")
                st.write(f"**Name:** {student_data['Student Name']}")
                st.write(f"**Original Section:** {student_data['Original Section']}")
                
                # Get attendance details
                attendance_data = get_student_attendance_details(student_data['Original Section'].replace("Original: ", ""), student)
                if attendance_data:
                    st.dataframe(attendance_data.sort_values('Date', ascending=False))
                    
                    if st.button("Download Student Report"):
                        csv = attendance_data.to_csv(index=False)
                        st.download_button(
                            label="Download CSV",
                            data=csv,
                            file_name=f"student_report_{student}_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                else:
                    st.info("No attendance records found")

def subject_analysis_page():
    """Page for subject-wise analysis"""
    st.subheader("Subject-wise Analysis")
    
    sections = get_sections()
    section = st.selectbox("Select Section", sections)
    
    if section:
        subjects = get_section_subjects(section)
        subject = st.selectbox("Select Subject", subjects)
        
        if subject:
            analysis_df = get_subject_analysis(section, subject)
            if not analysis_df.empty:
                st.write("### Subject Statistics")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    avg_attendance = analysis_df['Attendance %'].mean()
                    st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                with col2:
                    total_classes = analysis_df['Total Classes'].max()
                    st.metric("Total Classes", total_classes)
                with col3:
                    below_75 = len(analysis_df[analysis_df['Attendance %'] < 75])
                    st.metric("Students Below 75%", below_75)
                
                st.write("### Student-wise Analysis")
                st.dataframe(analysis_df)
                
                if st.button("Download Analysis"):
                    csv = analysis_df.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"subject_analysis_{section}_{subject}_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv"
                    )
            else:
                st.info("No data available for analysis")

def get_student_attendance_details(section, student_id):
    """Get detailed attendance records for a student"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=section)
        student_row = df[df['HT Number'] == student_id].iloc[0]
        
        attendance_data = []
        for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
            if pd.notna(student_row[period]) and student_row[period]:
                entries = str(student_row[period]).split('\n')
                for entry in entries:
                    if entry.strip():
                        try:
                            date, time, status, faculty, subject = entry.split('_')
                            attendance_data.append({
                                'Date': date,
                                'Time': time,
                                'Period': period,
                                'Status': status,
                                'Faculty': faculty,
                                'Subject': subject
                            })
                        except:
                            continue
        
        return pd.DataFrame(attendance_data)
    except Exception as e:
        st.error(f"Error getting attendance details: {str(e)}")
        return None

def get_subject_analysis(section, subject):
    """Get subject-wise attendance analysis"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        analysis = []
        
        for _, row in df.iterrows():
            present = 0
            total = 0
            
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if pd.notna(row[period]) and row[period]:
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        if subject in entry:
                            total += 1
                            if '_P_' in entry:
                                present += 1
            
            if total > 0:
                percentage = (present / total) * 100
            else:
                percentage = 0
                
            analysis.append({
                'HT Number': row['HT Number'],
                'Student Name': row['Student Name'],
                'Classes Attended': present,
                'Total Classes': total,
                'Attendance %': round(percentage, 2)
            })
        
        return pd.DataFrame(analysis)
    except Exception as e:
        st.error(f"Error in subject analysis: {str(e)}")
        return pd.DataFrame()


def workload_analysis_page():
    st.subheader("My Workload Analysis")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date", datetime.now().replace(day=1))
    with col2:
        to_date = st.date_input("To Date", datetime.now())
        
    # Get workload data
    total_periods, workload_df = get_faculty_workload(st.session_state.username)
    
    if not workload_df.empty:
        # Summary metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Classes", total_periods)
        with col2:
            unique_days = workload_df['Date'].nunique()
            st.metric("Days Engaged", unique_days)
        with col3:
            avg_classes = total_periods / max(unique_days, 1)
            st.metric("Daily Average", f"{avg_classes:.1f}")
        
        # Show section and subject breakdown
        st.subheader("Teaching Distribution")
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("##### Subject-wise Classes")
            subject_counts = workload_df['Subject'].value_counts().reset_index()
            subject_counts.columns = ['Subject', 'Classes']
            st.dataframe(subject_counts, hide_index=True)
            
        with col2:
            st.write("##### Section-wise Classes")
            section_counts = workload_df['Section'].value_counts().reset_index()
            section_counts.columns = ['Section', 'Classes']
            st.dataframe(section_counts, hide_index=True)
        
        # Detailed records
        st.subheader("Detailed Class Records")
        detailed_df = workload_df.sort_values('Date', ascending=False)
        st.dataframe(detailed_df, use_container_width=True)
        
        # Download option
        if st.button("Download Workload Report"):
            csv = detailed_df.to_csv(index=False)
            st.download_button(
                "Download CSV",
                data=csv,
                file_name=f"workload_report_{st.session_state.username}_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv"
            )
    else:
        st.info("No classes recorded in the selected date range")

# Add new function for faculty workload
def get_faculty_workload(username):
    """Calculate faculty workload"""
    try:
        all_sheets = pd.ExcelFile('attendance.xlsx').sheet_names
        section_sheets = [s for s in all_sheets if s.startswith('(O)')]
        
        total_periods = 0
        workload_data = []
        
        for sheet in section_sheets:
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if period in df.columns:
                    for _, row in df.iterrows():
                        if pd.notna(row[period]):
                            entries = str(row[period]).split('\n')
                            for entry in entries:
                                if username in entry:  # Check if faculty took this class
                                    date, time, status, faculty, subject = entry.split('_')
                                    workload_data.append({
                                        'Date': date,
                                        'Time': time,
                                        'Period': period,
                                        'Section': sheet.replace('(O)', ''),
                                        'Subject': subject
                                    })
                                    total_periods += 1
        
        return total_periods, pd.DataFrame(workload_data)
    except Exception as e:
        st.error(f"Error calculating workload: {str(e)}")
        return 0, pd.DataFrame()
def check_duplicate_attendance(section, period, date_str):
    """Check if attendance is already marked for given section, period and date"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        if period in df.columns:
            for value in df[period].dropna():
                for entry in str(value).split('\n'):
                    if entry.startswith(date_str):
                        return True
        return False
    except Exception as e:
        st.error(f"Error checking duplicate attendance: {str(e)}")
        return True



def get_attendance_stats(section, from_date=None, to_date=None):
    """Calculate attendance statistics in compact format"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        subjects = get_section_subjects(section)
        
        stats = []
        for _, row in df.iterrows():
            student_stats = {
                'HT Number': row['HT Number'],
                'Student Name': row['Student Name']
            }
            
            total_present = 0
            total_classes = 0
            
            for subject in subjects:
                present = 0
                total = 0
                
                for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                    if pd.notna(row[period]) and row[period]:
                        entries = str(row[period]).split('\n')
                        for entry in entries:
                            try:
                                if subject in entry:
                                    total += 1
                                    total_classes += 1
                                    if '_P_' in entry:
                                        present += 1
                                        total_present += 1
                            except:
                                continue
                
                # Add compact columns for each subject
                percentage = (present / total * 100) if total > 0 else 0
                student_stats[f"{subject} Present"] = present
                student_stats[f"{subject} Total"] = total
                student_stats[f"{subject} %"] = round(percentage, 2)
            
            # Calculate overall percentage
            overall_percentage = (total_present / total_classes * 100) if total_classes > 0 else 0
            student_stats.update({
                'Total Present': total_present,
                'Total Classes': total_classes,
                'Overall %': round(overall_percentage, 2)
            })
            
            stats.append(student_stats)
            
        return pd.DataFrame(stats)
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None

def view_statistics_page():
    """Page for viewing attendance statistics"""
    st.subheader("View Attendance Statistics")
    
    sections = [s.replace('(O)', '') for s in pd.ExcelFile('attendance.xlsx').sheet_names 
               if s.startswith('(O)')]
    section = st.selectbox("Select Section", sections)
    
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date")
    with col2:
        to_date = st.date_input("To Date")
    
    if section:
        stats_df = get_attendance_stats(
            section, 
            from_date.strftime('%Y-%m-%d') if from_date else None,
            to_date.strftime('%Y-%m-%d') if to_date else None
        )
        
        if stats_df is not None and not stats_df.empty:
            st.write("### Overall Statistics")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Students", len(stats_df))
            with col2:
                avg_attendance = stats_df['Overall %'].mean()
                st.metric("Average Attendance", f"{avg_attendance:.2f}%")
            with col3:
                below_75 = len(stats_df[stats_df['Overall %'] < 75])
                st.metric("Students Below 75%", below_75)
            
            # Add filter for percentage range
            min_attendance, max_attendance = st.select_slider(
                "Filter by Overall Attendance %",
                options=range(0, 101, 5),
                value=(0, 100)
            )
            
            # Filter dataframe
            filtered_df = stats_df[
                (stats_df['Overall %'] >= min_attendance) & 
                (stats_df['Overall %'] <= max_attendance)
            ]
            
            st.write("### Student-wise Statistics")
            st.dataframe(filtered_df)
            
            if st.button("Download Report"):
                csv = filtered_df.to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name=f"attendance_stats_{section}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
        else:
            st.info("No attendance records found for the selected criteria")
            

def get_column_width(col_name, values):
    try:
        max_length = max(
            max(len(str(val)) for val in values if val is not None),
            len(str(col_name))
        )
        return min(max_length * 10, 300)
    except:
        return 150

def admin_page():
    st.title("Admin Dashboard")
    
    with st.sidebar:
        st.header("Data Management")
        sheet = st.selectbox("Select Sheet", pd.ExcelFile('attendance.xlsx').sheet_names)
        
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    st.subheader(f"Edit {sheet}")
    
    try:
        # Read data
        df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
        
        # Convert all columns to string and handle special formatting
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].astype(float).astype(str)
            elif pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
            else:
                df[col] = df[col].astype(str)
        
        df = df.fillna('')
        
        # Remove empty rows
        df = df.loc[df.replace('', np.nan).notna().any(axis=1)].copy()
        df = df.reset_index(drop=True)
        
        # Calculate column heights based on content
        max_lines = {col: 1 for col in df.columns}
        for col in df.columns:
            for val in df[col]:
                if pd.notna(val) and val:
                    lines = str(val).count('\n') + 1
                    max_lines[col] = max(max_lines[col], lines)
    
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return

    # Configure columns with appropriate heights
    column_config = {}
    for col in df.columns:
        # Base height on content
        height = max(100, min(400, max_lines[col] * 35))
        
        # Different widths for different column types
        if col in ['Faculty Name', 'Credential', 'Student Name', 'HT Number']:
            width = 150
        else:
            width = 300
            
        column_config[str(col)] = st.column_config.TextColumn(
            str(col),
            width=width,
            help=f"Enter {col}",
            max_chars=None
        )

    # Add filters
    with st.expander("Show Filters", expanded=False):
        col1, col2, col3 = st.columns(3)
        
        with col1:
            search_text = st.text_input("Search", placeholder="Search in any column...")
            if search_text:
                mask = df.apply(lambda x: x.str.contains(search_text, case=False, na=False)).any(axis=1)
                df = df[mask]
        
        with col2:
            filter_column = st.selectbox("Filter by column", ["None"] + list(df.columns))
            if filter_column != "None":
                unique_values = sorted(df[filter_column].unique().tolist())
                selected_values = st.multiselect(
                    f"Select {filter_column} values",
                    unique_values,
                    default=unique_values
                )
                if selected_values:
                    df = df[df[filter_column].isin(selected_values)]
        
        with col3:
            sort_column = st.selectbox("Sort by", ["None"] + list(df.columns))
            if sort_column != "None":
                sort_order = st.radio("Sort order", ["Ascending", "Descending"])
                df = df.sort_values(by=sort_column, ascending=(sort_order == "Ascending"))

    # Add new row button
    if st.button("Add New Row"):
        new_row = pd.DataFrame([['' for _ in df.columns]], columns=df.columns)
        df = pd.concat([df, new_row], ignore_index=True)

    # Display editor with multiline support
    edited_df = st.data_editor(
        df,
        use_container_width=True,
        num_rows="dynamic",
        column_config=column_config,
        hide_index=True,
        height=min(800, len(df) * 50 + 100),  # Adjust height based on content
        key=f"editor_{sheet}"
    )

    # Remove empty rows before saving
    edited_df = edited_df.loc[edited_df.replace('', np.nan).notna().any(axis=1)].copy()

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Save Changes"):
            try:
                import os
                file_path = 'attendance.xlsx'
                
                # Check if file is writable
                if not os.access(file_path, os.W_OK):
                    st.error("Cannot write to file. Please check file permissions.")
                    return
                
                # Check if file is locked
                try:
                    with open(file_path, 'a') as f:
                        pass
                except IOError:
                    st.error("File is locked. Please close it in other programs and try again.")
                    return
                
                with pd.ExcelWriter(
                    file_path, 
                    mode='a', 
                    engine='openpyxl',
                    if_sheet_exists='overlay'
                ) as writer:
                    edited_df.to_excel(writer, sheet_name=sheet, index=False)
                    
                    worksheet = writer.sheets[sheet]
                    
                    # Format cells for multiline
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = openpyxl.styles.Alignment(
                                wrap_text=True,
                                vertical='top'
                            )
                            
                            # Set row height based on content
                            if cell.value:
                                lines = str(cell.value).count('\n') + 1
                                current_height = worksheet.row_dimensions[cell.row].height or 15
                                worksheet.row_dimensions[cell.row].height = max(current_height, lines * 15)
                    
                    # Set column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, max(len(line) for line in str(cell.value).split('\n')))
                        adjusted_width = min(50, max(12, max_length + 2))
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                st.success("Changes saved successfully!")
                st.rerun()
            except PermissionError:
                st.error("Cannot save changes. Please check if the file is open in another program.")
            except Exception as e:
                st.error(f"Error saving changes: {str(e)}")
                st.info("Try closing the Excel file if it's open in another program.")


def main():
    if 'logged_in' not in st.session_state:
        st.title("Login")
        
        # Show sample credentials
        with st.expander("View Demo Credentials"):
            st.info("""
            Admin Login:
            - Username: admin
            - Password: admin123
            
            Faculty Login:
            - Username: faculty1
            - Password: pass123
            """)
        
        login_type = st.radio("Select Login Type", ["Faculty", "Admin"])
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        
        if st.button("Login"):
            if check_login(username, password, login_type == "Admin"):
                st.session_state.logged_in = True
                st.session_state.is_admin = (login_type == "Admin")
                st.session_state.username = username
                st.rerun()
            else:
                st.error("Invalid credentials")
    else:
        if st.session_state.is_admin:
            admin_page()
        else:
            faculty_page()

if __name__ == "__main__":
    main()