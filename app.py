import streamlit as st
import pandas as pd
from datetime import datetime
import hashlib
import openpyxl
from openpyxl.styles import Alignment
import io
import numpy as np

# Single consolidated page config at the very start
st.set_page_config(
    page_title="Attendance Management System",
    layout="wide",
    initial_sidebar_state="collapsed",  # Better for mobile
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': "Attendance Management System"
    }
)

# Rest of your custom CSS for mobile-friendly styling
st.markdown("""
    <style>
        /* Mobile-friendly containers */
        .stApp {
            max-width: 100%;
            padding: 1rem;
        }
        
        /* Improved button styling */
        .stButton button {
            width: 100%;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            font-size: 1rem !important;
            font-weight: 500 !important;
            margin: 0.5rem 0 !important;
        }
        
        /* Card-like containers */
        .css-1r6slb0 {  /* Streamlit container class */
            background-color: white;
            padding: 1.5rem;
            border-radius: 15px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.05);
            margin: 0.5rem 0;
        }
        
        /* Responsive inputs */
        .stTextInput input, .stSelectbox select {
            width: 100%;
            padding: 0.8rem !important;
            border-radius: 10px !important;
            border: 1px solid #e0e0e0 !important;
        }
        
        /* Mobile-friendly metrics */
        .css-1xarl3l {  /* Metric container class */
            padding: 1rem !important;
            border-radius: 10px;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }
        
        /* Student list styling */
        .student-card {
            background-color: white;
            padding: 1rem;
            border-radius: 10px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            margin: 0.8rem 0;
        }
        
        /* Improved table responsiveness */
        .stDataFrame {
            overflow-x: auto;
        }
        
        /* Better spacing for mobile */
        @media (max-width: 768px) {
            .stApp {
                padding: 0.5rem;
            }
            
            .row-widget {
                margin: 0.5rem 0 !important;
            }
            
            /* Stack columns on mobile */
            .css-12w0qpk {
                flex-direction: column;
            }
            
            .css-1d391kg {
                width: 100% !important;
            }
        }
        
        /* Floating action button for submit */
        .submit-button {
            position: fixed;
            bottom: 20px;
            right: 20px;
            z-index: 999;
            width: auto !important;
            padding: 1rem 2rem !important;
            box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        }
        .attendance-status {
            background: #f8fafc;
            border-left: 1px solid #e2e8f0;
            padding: 0.75rem;
            min-width: 120px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
    </style>
""", unsafe_allow_html=True)

# Admin credentials
ADMIN_CREDENTIALS = {
    "a": hashlib.sha256("a".encode()).hexdigest()
}

# Rest of your code remains the same...


def mark_attendance_page():
    """Enhanced mobile-friendly attendance marking page"""
    section = st.session_state.sections[0] if st.session_state.sections else None
    subject = st.session_state.subject
    period = st.session_state.period

    if section and subject and period:
        df_students = get_student_data(section)
        if df_students is not None:
            attendance_data = {}
            
            # Session info card
            st.markdown(f"""
                <div style='background: linear-gradient(135deg, #6B46C1 0%, #805AD5 100%);
                          color: white;
                          padding: 1.2rem;
                          border-radius: 15px;
                          margin-bottom: 1.5rem;
                          box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);'>
                    <h3 style='margin: 0; font-size: 1.2rem; font-weight: 600;'>Current Session</h3>
                    <div style='margin-top: 1rem;'>
                        <div style='display: flex; align-items: center; margin: 0.5rem 0;'>
                            <span style='width: 24px; text-align: center; margin-right: 8px;'>📚</span>
                            <span style='font-size: 1rem;'>{section}</span>
                        </div>
                        <div style='display: flex; align-items: center; margin: 0.5rem 0;'>
                            <span style='width: 24px; text-align: center; margin-right: 8px;'>📖</span>
                            <span style='font-size: 1rem;'>{subject}</span>
                        </div>
                        <div style='display: flex; align-items: center; margin: 0.5rem 0;'>
                            <span style='width: 24px; text-align: center; margin-right: 8px;'>⏰</span>
                            <span style='font-size: 1rem;'>Period {period}</span>
                        </div>
                    </div>
                </div>
            """, unsafe_allow_html=True)
            
            # Quick action buttons
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✓ Mark All Present", use_container_width=True, type="primary"):
                    st.session_state.select_all = True
            with col2:
                if st.button("✗ Mark All Absent", use_container_width=True):
                    st.session_state.select_all = False
            
            # Student list with clean cards
            for idx, student in df_students.iterrows():
                # Create a container for each student card
                with st.container():
                    col1, col2 = st.columns([7,3])
                    
                    # Student info column
                    with col1:
                        st.markdown(f"""
                            <div style="padding-top: 0.5rem;">
                                <div style="font-size: 1rem; font-weight: 500; color: white; margin-bottom: 0.2rem;">
                                    {student['Student Name']}
                                </div>
                                <div style="font-size: 1rem; color: white; margin-bottom: 0.2rem;">
                                    {student['HT Number']}
                                </div>
                                <div style="font-size: 0.8rem; color: #888;">
                                    {student['Original Section']}
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                    
                    # Attendance checkbox column
                    with col2:
                        status = st.checkbox(
                            "Present",
                            key=student['HT Number'],
                            value=getattr(st.session_state, 'select_all', True)
                        )
                        
                    attendance_data[student['HT Number']] = {
                        'status': 'P' if status else 'A',
                        'original_section': student['Original Section']
                    }
                    
                    # Add a divider between cards
                    st.markdown("<hr style='margin: 0.5rem 0; border: none; border-top: 1px solid #eee;'>", unsafe_allow_html=True)
            
            # Submit button with margin space
            st.markdown("<div style='height: 60px;'></div>", unsafe_allow_html=True)
            
            if st.button("📝 Submit Attendance", type="primary", key="submit_attendance", use_container_width=True):
                success, unsuccessful_records = mark_attendance(
                    section, period, attendance_data,
                    st.session_state.username, subject
                )
                
                if unsuccessful_records:
                    st.info(f"✅ Recorded {len(attendance_data) - len(unsuccessful_records)} students")
                    st.warning("⚠️ Issues found:")
                    for record in unsuccessful_records:
                        st.markdown(f"""
                            <div style='background: #FFF3CD;
                                      padding: 1rem;
                                      border-radius: 8px;
                                      margin: 0.5rem 0;
                                      border: 1px solid #FFE69C;'>
                                <div style='font-weight: 500; color: #664D03;'>
                                    {record['Student Name']} ({record['HT Number']})
                                </div>
                                <div style='color: #997404; font-size: 0.9rem; margin-top: 0.3rem;'>
                                    {record['Reason']}
                                </div>
                            </div>
                        """, unsafe_allow_html=True)
                else:
                    st.success(f"✅ Successfully recorded all {len(attendance_data)} students")

def faculty_page():
    # Get faculty full name from session state
    faculty_name = st.session_state.faculty_name
    
    # Use faculty name for welcome message
    st.title(f"Welcome, {faculty_name}")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio(
            "Select", 
            ["Mark Attendance", "View Statistics", "Student Reports", 
             "Subject Analysis", "My Workload"]
        )

        if page == "Mark Attendance":
            # Period selection
            st.session_state.period = st.selectbox(
                "Select Period",
                options=[''] + ['P1', 'P2', 'P3', 'P4', 'P5', 'P6'],
                key="period_select"
            )
            
            # Section selection - only manipulated sections for attendance
            # Changed from multiselect to selectbox for single selection
            sections = get_sections(for_attendance=True)
            selected_section = st.selectbox(
                "Select Section",
                options=[''] + sections,
                key="section_select"
            )
            # Store the selected section in session state as a list with single item
            st.session_state.sections = [selected_section] if selected_section else []
            
            # Show available subjects for selected section
            if st.session_state.sections:
                all_subjects = []
                for section in st.session_state.sections:
                    subjects = get_section_subjects(section)
                    all_subjects.extend(subjects)
                # Remove duplicates and sort
                unique_subjects = sorted(list(set(all_subjects)))
                
                st.session_state.subject = st.selectbox(
                    "Select Subject",
                    options=unique_subjects if unique_subjects else [''],
                    help="Select the subject being taught",
                    key="subject_select"
                )
        
        if st.button("Logout", key="logout_button"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    # Page content based on selection
    if page == "Mark Attendance":
        if not (st.session_state.get('period') and st.session_state.get('sections') 
                and st.session_state.get('subject')):
            st.info("Please select Period, Section, and Subject from the sidebar")
            return
        mark_attendance_page()
    elif page == "View Statistics":
        view_statistics_page()
    elif page == "Student Reports":
        student_reports_page()
    elif page == "Subject Analysis":
        subject_analysis_page()
    else:  # My Workload
        workload_analysis_page()


# Updated mark_attendance function to return both success status and unsuccessful records
def update_faculty_log(faculty_name, section, period, subject):
    """Update faculty attendance log with month-wise organization"""
    try:
        # Read faculty sheet
        df = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        
        # Get current month-year
        current_date = datetime.now()
        month_year = current_date.strftime('%b%Y')
        
        # Create log entry in the format: DD/MM/YYYY_HH:MM[AM/PM]_[Period]_[Subject]_[Section]
        log_entry = f"{current_date.strftime('%d/%m/%Y_%I:%M%p')}_{period}_{subject}_{section}"
        
        # Check if month-year column exists, if not create it
        if month_year not in df.columns:
            # Get existing columns
            existing_cols = list(df.columns)
            # Find the position after 'Password' column
            password_idx = existing_cols.index('Password')
            # Insert new column after Password
            df.insert(password_idx + 1, month_year, '')
        
        # Update the log for the faculty
        faculty_mask = df['Faculty Name'] == faculty_name
        if faculty_mask.any():
            current_log = df.loc[faculty_mask, month_year].iloc[0]
            # Add new entry with proper newline handling
            if pd.notna(current_log) and str(current_log).strip():
                new_log = f"{current_log}\n{log_entry}"
            else:
                new_log = log_entry
            df.loc[faculty_mask, month_year] = new_log
        
        # Save the updated sheet while preserving column order
        with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Faculty', index=False)
            
            # Format the worksheet
            worksheet = writer.sheets['Faculty']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Set column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                adjusted_width = min(50, max(12, max_length + 2))
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        return True
    except Exception as e:
        st.error(f"Error updating faculty log: {str(e)}")
        return False




def get_student_data(section):
    """Get student data for a manipulated section"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=section)
        students_df = df[['HT Number', 'Student Name', 'Original Section']].fillna('')
        # Format original section display without duplication
        students_df['Original Section'] = students_df['Original Section'].apply(
            lambda x: f"Original: {x}" if not x.startswith("Original:") else x
        )
        return students_df
    except Exception as e:
        st.error(f"Error getting student data: {str(e)}")
        return None




def get_student_attendance_details(section, student_id):
    """Get detailed attendance records for a student"""
    try:
        # Remove (O) prefix if present
        section = section.replace("(O)", "").strip()
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        student_row = df[df['HT Number'] == student_id]
        
        if student_row.empty:
            return None
            
        student_row = student_row.iloc[0]
        attendance_data = []
        
        for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
            if pd.notna(student_row[period]) and student_row[period]:
                entries = str(student_row[period]).split('\n')
                for entry in entries:
                    if entry.strip():
                        try:
                            date, time, status, faculty, subject = entry.split('_')
                            attendance_data.append({
                                'Date': date,
                                'Time': time,
                                'Period': period,
                                'Status': status,
                                'Faculty': faculty,
                                'Subject': subject
                            })
                        except ValueError:
                            continue  # Skip malformed entries
        
        if not attendance_data:  # If no attendance records found
            return pd.DataFrame()  # Return empty DataFrame
            
        return pd.DataFrame(attendance_data)
    except Exception as e:
        st.error(f"Error getting attendance details: {str(e)}")
        return None





def check_duplicate_attendance(section, period, date_str):
    """Check if attendance is already marked for given section, period and date"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        if period in df.columns:
            for value in df[period].dropna():
                for entry in str(value).split('\n'):
                    if entry.startswith(date_str):
                        return True
        return False
    except Exception as e:
        st.error(f"Error checking duplicate attendance: {str(e)}")
        return True




def get_sections(for_attendance=False):
    """
    Get sections based on context
    for_attendance=True: Returns only manipulated sections (without O prefix)
    for_attendance=False: Returns original sections with (O) prefix intact
    """
    try:
        all_sheets = pd.ExcelFile('attendance.xlsx').sheet_names
        if for_attendance:
            # For attendance marking: Only show manipulated sections
            return [s for s in all_sheets if not s.startswith('(O)') 
                   and s not in ['Faculty', 'Section-Subject-Mapping']]
        else:
            # For other features: Show original sections with (O) prefix
            return [s for s in all_sheets 
                   if s.startswith('(O)') and s not in ['Faculty', 'Section-Subject-Mapping']]
    except Exception as e:
        st.error(f"Error getting sections: {str(e)}")
        return []

def student_reports_page():
    """Page for individual student reports with multi-select sections"""
    st.subheader("Individual Student Reports")
    
    # Use original sections with (O) prefix for reports
    sections = get_sections(for_attendance=False)
    selected_sections = st.multiselect("Select Sections", options=sections)
    
    if selected_sections:
        all_students = []
        for section in selected_sections:
            try:
                df = pd.read_excel('attendance.xlsx', sheet_name=section)
                df['Section'] = section  # Add section column
                all_students.append(df[['HT Number', 'Student Name', 'Section']])
            except Exception as e:
                st.error(f"Error reading section {section}: {str(e)}")
                continue
        
        if all_students:
            df_students = pd.concat(all_students, ignore_index=True)
            student = st.selectbox(
                "Select Student", 
                df_students['HT Number'].tolist(),
                format_func=lambda x: f"{x} - {df_students[df_students['HT Number']==x]['Student Name'].iloc[0]} ({df_students[df_students['HT Number']==x]['Section'].iloc[0]})"
            )
            
            if student:
                student_data = df_students[df_students['HT Number'] == student].iloc[0]
                
                st.write(f"### Attendance Report for {student}")
                st.write(f"**Name:** {student_data['Student Name']}")
                st.write(f"**Section:** {student_data['Section']}")
                
                # Get attendance details
                attendance_data = get_student_attendance_details(student_data['Section'], student)
                
                if attendance_data is not None and not attendance_data.empty:
                    st.dataframe(attendance_data.sort_values('Date', ascending=False))
                    
                    if st.button("Download Student Report"):
                        csv = attendance_data.to_csv(index=False)
                        st.download_button(
                            label="Download CSV",
                            data=csv,
                            file_name=f"student_report_{student}_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                else:
                    st.info("No attendance records found")

def get_attendance_stats(section, from_date=None, to_date=None):
    """Calculate attendance statistics with attended/conducted format"""
    try:
        # Use the original section name with (O) prefix if not present
        if not section.startswith('(O)'):
            section = f"(O){section}"

        df = pd.read_excel('attendance.xlsx', sheet_name=section)
        subjects = get_section_subjects(section)
        if not subjects:
            st.error(f"No subjects found for section {section}")
            return None
        
        # Convert date inputs to datetime.date objects if they're strings
        if isinstance(from_date, str):
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        if isinstance(to_date, str):
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        
        stats = []
        for _, row in df.iterrows():
            if pd.isna(row['HT Number']) or pd.isna(row['Student Name']):
                continue
                
            student_stats = {
                'HT Number': str(row['HT Number']),
                'Student Name': str(row['Student Name'])
            }
            
            total_present = 0
            total_classes = 0
            
            # Calculate for each mapped subject
            for subject in subjects:
                present = 0
                total = 0
                
                for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                    if pd.notna(row[period]) and str(row[period]).strip():
                        entries = str(row[period]).split('\n')
                        for entry in entries:
                            try:
                                parts = entry.split('_')
                                if len(parts) >= 5:  # Ensure we have all required parts
                                    date_str = parts[0]
                                    entry_date = datetime.strptime(date_str, '%d/%m/%Y').date()
                                    
                                    # Check if entry is within date range
                                    if from_date and to_date:
                                        if not (from_date <= entry_date <= to_date):
                                            continue
                                    
                                    # Check if this entry is for the current subject
                                    subject_name = parts[4]  # The subject name should be in this position
                                    if subject in subject_name:
                                        total += 1
                                        total_classes += 1
                                        if '_P_' in entry:
                                            present += 1
                                            total_present += 1
                            except (ValueError, IndexError) as e:
                                st.error(f"Error processing entry: {entry}, Error: {str(e)}")
                                continue
                
                # Only add subject column if there are classes for this subject
                if total > 0:
                    student_stats[f"{subject}\n(Attended/Conducted)"] = f"{present}/{total}"
            
            # Only add student stats if they attended any classes
            if total_classes > 0:
                # Add total attended/conducted
                student_stats[f"Total\n(Attended/Conducted)"] = f"{total_present}/{total_classes}"
                
                # Calculate overall percentage
                overall_percentage = (total_present / total_classes * 100) if total_classes > 0 else 0
                student_stats['Overall %'] = round(overall_percentage, 2)
                
                stats.append(student_stats)
        
        if not stats:
            return pd.DataFrame()
            
        # Convert to DataFrame
        stats_df = pd.DataFrame(stats)
        
        # Ensure correct column order
        base_columns = ['HT Number', 'Student Name']
        subject_columns = [col for col in stats_df.columns 
                         if '(Attended/Conducted)' in col and 'Total' not in col]
        total_column = [col for col in stats_df.columns if col.startswith('Total')]
        percentage_column = ['Overall %']
        
        # Reorder columns
        ordered_columns = base_columns + subject_columns + total_column + percentage_column
        stats_df = stats_df[ordered_columns]
        
        return stats_df
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return pd.DataFrame()


def view_statistics_page():
    """Page for viewing attendance statistics with multi-select sections"""
    st.subheader("View Attendance Statistics")
    
    # Get original sections with (O) prefix
    sections = get_sections(for_attendance=False)
    selected_sections = st.multiselect("Select Sections", options=sections)
    
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date")
    with col2:
        to_date = st.date_input("To Date")
    
    if selected_sections:
        all_stats = []
        for section in selected_sections:
            stats_df = get_attendance_stats(
                section, 
                from_date,
                to_date
            )
            if stats_df is not None and not stats_df.empty:
                stats_df['Section'] = section
                all_stats.append(stats_df)
        
        if all_stats:
            combined_stats = pd.concat(all_stats, ignore_index=True)
            
            st.write("### Overall Statistics")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Students", len(combined_stats))
            with col2:
                avg_attendance = combined_stats['Overall %'].mean()
                st.metric("Average Attendance", f"{avg_attendance:.2f}%")
            with col3:
                below_75 = len(combined_stats[combined_stats['Overall %'] < 75])
                st.metric("Students Below 75%", below_75)
            
            # Configure column display
            column_config = {
                'HT Number': st.column_config.TextColumn('HT Number', width=120),
                'Student Name': st.column_config.TextColumn('Student Name', width=180),
                'Section': st.column_config.TextColumn('Section', width=150),
                'Overall %': st.column_config.NumberColumn(
                    'Overall %',
                    format="%.2f%%",
                    width=100
                )
            }
            
            # Add configuration for subject and total columns
            for col in combined_stats.columns:
                if 'Attended/Conducted' in col:
                    column_config[col] = st.column_config.TextColumn(
                        col,
                        width=150
                    )
            
            st.write("### Student-wise Statistics")
            st.dataframe(
                combined_stats,
                column_config=column_config,
                use_container_width=True,
                hide_index=True
            )
            
            if st.button("Download Report"):
                csv = combined_stats.to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name=f"attendance_stats_combined_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
        else:
            st.info("No attendance records found for the selected criteria")

def get_column_width(col_name, values):
    try:
        max_length = max(
            max(len(str(val)) for val in values if val is not None),
            len(str(col_name))
        )
        return min(max_length * 10, 300)
    except:
        return 150



def validate_upload_data(df, sheet_name):
    """Validate uploaded data against expected format"""
    template_df = create_template_df(sheet_name)
    return all(col in df.columns for col in template_df.columns)

def show_data_editor(sheet):
    """Show the data editor component with improved layout"""
    try:
        if sheet == 'Faculty':
            # Modified to load all columns for Faculty sheet
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            df = df.fillna('')
            
            # Configure columns with appropriate widths
            column_config = {}
            for col in df.columns:
                width = 150 if col in ['Faculty Name', 'Username', 'Password'] else 300
                column_config[str(col)] = st.column_config.TextColumn(
                    col,
                    width=width,
                    help=f"Enter {col}",
                    max_chars=None
                )
        else:
            # Regular handling for other sheets
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            df = df.convert_dtypes().fillna('')
            
            # Configure columns
            column_config = {}
            for col in df.columns:
                width = 150 if col in ['HT Number', 'Student Name'] else 300
                column_config[str(col)] = st.column_config.TextColumn(
                    col,
                    width=width,
                    help=f"Enter {col}",
                    max_chars=None
                )
        
        # Get actual number of rows
        data_rows = len(df)
        
        # Display editor with minimal extra rows
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            column_config=column_config,
            hide_index=True,
            height=min(600, (data_rows + 2) * 35)
        )
        
        # Create columns for buttons
        col1, col2, col3 = st.columns([1, 1, 2])
        
        with col1:
            if st.button("Save Changes"):
                try:
                    with pd.ExcelWriter('attendance.xlsx', 
                                      mode='a', 
                                      if_sheet_exists='overlay',
                                      engine='openpyxl') as writer:
                        edited_df.to_excel(writer, sheet_name=sheet, index=False)
                        
                        # Format worksheet
                        worksheet = writer.sheets[sheet]
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        # Set column widths
                        for column in worksheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            adjusted_width = min(50, max(12, max_length + 2))
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    st.success("Changes saved successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error saving changes: {str(e)}")

        with col2:
            # Download current sheet button
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                edited_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
            
            st.download_button(
                label="Download Sheet",
                data=buffer.getvalue(),
                file_name=f"{sheet}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
                
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")


def get_faculty_data(sheet):
    """Get faculty data with proper type conversion"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
        # Convert all columns to string type and handle dates
        for col in df.columns:
            df[col] = df[col].apply(lambda x: str(x) if pd.notnull(x) else '')
        return df
    except Exception as e:
        st.error(f"Error getting faculty data: {str(e)}")
        return None


def get_all_faculty_workload(from_date=None, to_date=None):
    """Get workload statistics for all faculty members"""
    try:
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        faculty_stats = []
        
        for _, faculty_row in df_faculty.iterrows():
            faculty_name = faculty_row['Faculty Name']
            workload_data = []
            
            # Process each month column
            for col in faculty_row.index:
                if col not in ['Faculty Name', 'Password']:
                    entries = str(faculty_row[col]).split('\n') if pd.notna(faculty_row[col]) else []
                    
                    for entry in entries:
                        if entry.strip():
                            try:
                                parts = entry.strip().split('_')
                                if len(parts) >= 5:
                                    date_str = parts[0]
                                    time_str = parts[1]
                                    period = parts[2]
                                    subject = parts[3]
                                    section = parts[4]
                                    
                                    date_obj = pd.to_datetime(date_str, format='%d/%m/%Y')
                                    
                                    # Apply date filter if provided
                                    if from_date and to_date:
                                        if not (pd.to_datetime(from_date) <= date_obj <= pd.to_datetime(to_date)):
                                            continue
                                    
                                    workload_data.append({
                                        'Date': date_str,
                                        'Time': time_str,
                                        'Period': period,
                                        'Subject': subject,
                                        'Section': section
                                    })
                            except Exception:
                                continue
            
            # Calculate statistics for this faculty
            if workload_data:
                df_workload = pd.DataFrame(workload_data)
                unique_days = len(df_workload['Date'].unique())
                unique_subjects = len(df_workload['Subject'].unique())
                unique_sections = len(df_workload['Section'].unique())
                total_classes = len(workload_data)
                
                # Get subject and section distribution
                subject_dist = df_workload['Subject'].value_counts().to_dict()
                section_dist = df_workload['Section'].value_counts().to_dict()
                
                faculty_stats.append({
                    'Faculty Name': faculty_name,
                    'Total Classes': total_classes,
                    'Days Engaged': unique_days,
                    'Daily Average': round(total_classes / max(unique_days, 1), 2),
                    'Unique Subjects': unique_subjects,
                    'Unique Sections': unique_sections,
                    'Subject Distribution': subject_dist,
                    'Section Distribution': section_dist,
                    'Detailed Records': df_workload
                })
        
        return faculty_stats
    except Exception as e:
        st.error(f"Error calculating faculty workload: {str(e)}")
        return []

def show_faculty_workload_admin():
    """Display faculty workload overview for admin"""
    st.subheader("Faculty Workload Overview")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input(
            "From Date",
            datetime.now().replace(day=1),
            format="YYYY/MM/DD"
        )
    with col2:
        to_date = st.date_input(
            "To Date",
            datetime.now(),
            format="YYYY/MM/DD"
        )
    
    # Get faculty workload data
    faculty_stats = get_all_faculty_workload(from_date, to_date)
    
    if faculty_stats:
        # Summary metrics
        st.write("### Overall Statistics")
        
        # Calculate overall metrics
        total_faculty = len(faculty_stats)
        total_classes = sum(stat['Total Classes'] for stat in faculty_stats)
        avg_classes_per_faculty = total_classes / total_faculty
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Faculty", total_faculty)
        with col2:
            st.metric("Total Classes Conducted", total_classes)
        with col3:
            st.metric("Avg Classes per Faculty", f"{avg_classes_per_faculty:.1f}")
        
        # Faculty-wise breakdown
        st.write("### Faculty-wise Analysis")
        
        # Create summary DataFrame
        summary_data = []
        for stat in faculty_stats:
            summary_data.append({
                'Faculty Name': stat['Faculty Name'],
                'Total Classes': stat['Total Classes'],
                'Days Engaged': stat['Days Engaged'],
                'Daily Average': stat['Daily Average'],
                'Subjects Handled': stat['Unique Subjects'],
                'Sections Handled': stat['Unique Sections']
            })
        
        summary_df = pd.DataFrame(summary_data)
        st.dataframe(
            summary_df,
            column_config={
                'Faculty Name': st.column_config.TextColumn('Faculty Name', width=150),
                'Total Classes': st.column_config.NumberColumn('Total Classes', width=100),
                'Days Engaged': st.column_config.NumberColumn('Days Engaged', width=100),
                'Daily Average': st.column_config.NumberColumn('Daily Average', format="%.2f", width=100),
                'Subjects Handled': st.column_config.NumberColumn('Subjects', width=100),
                'Sections Handled': st.column_config.NumberColumn('Sections', width=100)
            },
            hide_index=True
        )
        
        # Detailed faculty records
        st.write("### Detailed Faculty Records")
        for stat in faculty_stats:
            with st.expander(f"📊 {stat['Faculty Name']}"):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.write("##### Subject Distribution")
                    subject_df = pd.DataFrame([
                        {'Subject': subject, 'Classes': count}
                        for subject, count in stat['Subject Distribution'].items()
                    ])
                    st.dataframe(subject_df, hide_index=True)
                
                with col2:
                    st.write("##### Section Distribution")
                    section_df = pd.DataFrame([
                        {'Section': section, 'Classes': count}
                        for section, count in stat['Section Distribution'].items()
                    ])
                    st.dataframe(section_df, hide_index=True)
                
                st.write("##### Detailed Class Records")
                detailed_df = stat['Detailed Records'].sort_values('Date', ascending=False)
                st.dataframe(
                    detailed_df,
                    column_config={
                        'Date': st.column_config.TextColumn('Date', width=100),
                        'Time': st.column_config.TextColumn('Time', width=100),
                        'Period': st.column_config.TextColumn('Period', width=80),
                        'Subject': st.column_config.TextColumn('Subject', width=150),
                        'Section': st.column_config.TextColumn('Section', width=150)
                    },
                    hide_index=True
                )
                
                # Download option for individual faculty
                csv = detailed_df.to_csv(index=False)
                st.download_button(
                    label=f"Download {stat['Faculty Name']}'s Records",
                    data=csv,
                    file_name=f"workload_{stat['Faculty Name']}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
        
        # Download complete report
        st.write("### Download Options")
        if st.button("Download Complete Report"):
            # Create Excel file with multiple sheets
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Write summary sheet
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Write individual faculty sheets
                for stat in faculty_stats:
                    stat['Detailed Records'].to_excel(
                        writer, 
                        sheet_name=f"{stat['Faculty Name'][:30]}", 
                        index=False
                    )
            
            st.download_button(
                label="📥 Download Complete Workload Report",
                data=buffer.getvalue(),
                file_name=f"faculty_workload_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.info("No faculty workload data available for the selected date range")

# Update admin_page function to include faculty workload
def admin_page():
    st.title("Admin Dashboard")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio(
            "Select",
            ["Data Management", "Faculty Workload"]
        )
    
    if page == "Data Management":
        try:
            with st.sidebar:
                st.header("Data Management")
                sheet = st.selectbox("Select Sheet", pd.ExcelFile('attendance.xlsx').sheet_names)
                
                st.write("### Download Options")
                
                # Upload complete workbook button
                uploaded_workbook = st.file_uploader("Upload Complete Workbook", type=['xlsx'])
                if uploaded_workbook is not None:
                    try:
                        xls = pd.ExcelFile(uploaded_workbook)
                        st.write("Sheets found in workbook:", xls.sheet_names)
                        
                        if st.button("Confirm Upload"):
                            with pd.ExcelWriter('attendance.xlsx', 
                                              mode='a', 
                                              if_sheet_exists='overlay') as writer:
                                for sheet_name in xls.sheet_names:
                                    df = pd.read_excel(uploaded_workbook, sheet_name=sheet_name)
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                            st.success("Workbook uploaded successfully!")
                            st.rerun()
                    except Exception as e:
                        st.error(f"Error uploading workbook: {str(e)}")
                
                # Download complete workbook
                if st.button("Download Complete Workbook"):
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        xls = pd.ExcelFile('attendance.xlsx')
                        for sheet_name in xls.sheet_names:
                            df = pd.read_excel('attendance.xlsx', sheet_name=sheet_name, dtype=str)
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            worksheet = writer.sheets[sheet_name]
                            for column in worksheet.columns:
                                max_length = max(len(str(cell.value or '')) for cell in column)
                                worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                    
                    st.download_button(
                        label="📥 Download Complete Workbook",
                        data=buffer.getvalue(),
                        file_name="attendance_workbook.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                
            # Add tabs for different operations
            tab1, tab2 = st.tabs(["Edit Data", "Bulk Upload"])
            
            with tab1:
                st.subheader(f"Edit {sheet}")
                show_data_editor(sheet)
            
            with tab2:
                st.subheader("Bulk Upload")
                show_bulk_upload(sheet)
                
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
    
    else:  # Faculty Workload page
        show_faculty_workload_admin()
    
    # Logout button at the bottom of sidebar
    with st.sidebar:
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

def show_bulk_upload(sheet):
    """Show bulk upload interface"""
    # Download template button
    if st.button("Download Template"):
        template_df = create_template_df(sheet)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            template_df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
        st.download_button(
            label="📥 Download Template Excel",
            data=buffer.getvalue(),
            file_name=f"{sheet}_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Upload file
    uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx'])
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            st.write("Preview of uploaded data:")
            st.dataframe(df)
            
            if st.button("Confirm Upload"):
                if validate_upload_data(df, sheet):
                    with pd.ExcelWriter('attendance.xlsx', 
                                      mode='a',
                                      if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, sheet_name=sheet, index=False)
                    st.success("Data uploaded successfully!")
                    st.rerun()
                else:
                    st.error("Invalid data format. Please use the template.")
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")


def create_template_df(sheet_name):
    """Create template DataFrame based on sheet type with updated Faculty template"""
    if sheet_name == 'Faculty':
        # Get current month-year
        current_month = datetime.now().strftime('%b%Y')
        return pd.DataFrame(columns=['Faculty Name', 'Username', 'Password', current_month])
    elif sheet_name == 'Section-Subject-Mapping':
        return pd.DataFrame(columns=['Section', 'Subject Names'])
    elif sheet_name.startswith('(O)'):
        return pd.DataFrame(columns=['HT Number', 'Student Name', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6'])
    else:  # Manipulated sections
        return pd.DataFrame(columns=['HT Number', 'Student Name', 'Original Section'])

def get_faculty_id(faculty_name):
    """Extract RVIT ID from faculty name"""
    if '(' in faculty_name and ')' in faculty_name:
        return faculty_name[faculty_name.index('(')+1:faculty_name.index(')')]
    return None

def check_login(username, password, is_admin=False):
    """Verify login credentials"""
    try:
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        if is_admin:
            # Check admin credentials
            return any((df_faculty['Username'] == username) & 
                      (df_faculty['Password'] == password) &
                      (df_faculty['Faculty Name'].str.contains('admin', case=False)))
        else:
            # Check faculty credentials
            return any((df_faculty['Username'] == username) & 
                      (df_faculty['Password'] == password))
    except Exception as e:
        st.error(f"Login error: {str(e)}")
        return False



def get_section_subjects(section, for_subject_analysis=False):
    """
    Get subjects mapped to a specific section
    for_subject_analysis=True: Use manipulated section name
    for_subject_analysis=False: Use original section name with (O) prefix
    """
    try:
        df_mapping = pd.read_excel('attendance.xlsx', sheet_name='Section-Subject-Mapping')
        
        # For subject analysis, use the section name as-is (manipulated section)
        if for_subject_analysis:
            matching_rows = df_mapping[df_mapping['Section'].str.strip() == section.strip()]
        else:
            # For other features, handle original sections with (O) prefix
            if section.startswith('(O)'):
                clean_section = section.replace('(O)', '').strip()
                matching_rows = df_mapping[df_mapping['Section'].str.strip() == clean_section]
            else:
                matching_rows = df_mapping[df_mapping['Section'].str.strip() == section.strip()]
        
        if matching_rows.empty:
            st.error(f"No subject mapping found for section: {section}")
            return []
            
        # Get subjects from the first matching row
        subjects_str = str(matching_rows.iloc[0]['Subject Names'])
        # Split subjects and clean up
        subjects = [s.strip() for s in subjects_str.split('\n') if s.strip()]
        return subjects
    except Exception as e:
        st.error(f"Error getting subjects: {str(e)}")
        return []

def get_subject_analysis(section, subject):
    """Get subject-wise attendance analysis"""
    try:
        # For original section name, check if (O) prefix exists
        sheet_name = section if section.startswith('(O)') else f'(O){section}'
        df = pd.read_excel('attendance.xlsx', sheet_name=sheet_name)
        analysis = []
        
        for _, row in df.iterrows():
            if pd.isna(row['HT Number']) or pd.isna(row['Student Name']):
                continue
                
            present = 0
            total = 0
            
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if pd.notna(row[period]) and str(row[period]).strip():
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        if subject in entry:
                            total += 1
                            if '_P_' in entry:
                                present += 1
            
            if total > 0:
                percentage = (present / total) * 100
            else:
                percentage = 0
                
            analysis.append({
                'HT Number': str(row['HT Number']),
                'Student Name': str(row['Student Name']),
                'Classes Attended': present,
                'Total Classes': total,
                'Attendance %': round(percentage, 2)
            })
        
        return pd.DataFrame(analysis)
    except Exception as e:
        st.error(f"Error in subject analysis: {str(e)}")
        return pd.DataFrame()

def subject_analysis_page():
    """Page for subject-wise analysis"""
    st.subheader("Subject-wise Analysis")
    
    # Use manipulated sections (without O prefix) for subject analysis
    sections = get_sections(for_attendance=True)  # This gets sections without (O) prefix
    section = st.selectbox("Select Section", sections)
    
    if section:
        subjects = get_section_subjects(section, for_subject_analysis=True)
        if subjects:  # Only show subject selection if subjects were found
            subject = st.selectbox("Select Subject", subjects)
            
            if subject:
                # Find the corresponding original section name
                original_section = f"(O){section}"
                try:
                    analysis_df = get_subject_analysis(original_section, subject)
                    if not analysis_df.empty:
                        st.write("### Subject Statistics")
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            avg_attendance = analysis_df['Attendance %'].mean()
                            st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                        with col2:
                            total_classes = analysis_df['Total Classes'].max()
                            st.metric("Total Classes", total_classes)
                        with col3:
                            below_75 = len(analysis_df[analysis_df['Attendance %'] < 75])
                            st.metric("Students Below 75%", below_75)
                        
                        st.write("### Student-wise Analysis")
                        st.dataframe(
                            analysis_df.sort_values('Attendance %', ascending=False),
                            column_config={
                                'HT Number': st.column_config.TextColumn('HT Number', width=120),
                                'Student Name': st.column_config.TextColumn('Student Name', width=180),
                                'Classes Attended': st.column_config.NumberColumn('Classes Attended', width=130),
                                'Total Classes': st.column_config.NumberColumn('Total Classes', width=120),
                                'Attendance %': st.column_config.NumberColumn('Attendance %', format="%.2f%%", width=120)
                            },
                            hide_index=True,
                            use_container_width=True
                        )
                        
                        if st.button("Download Analysis"):
                            csv = analysis_df.to_csv(index=False)
                            st.download_button(
                                label="Download CSV",
                                data=csv,
                                file_name=f"subject_analysis_{section}_{subject}_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv"
                            )
                    else:
                        st.info(f"No attendance records found for {subject} in {section}")
                except Exception as e:
                    st.error(f"Error accessing attendance data: {str(e)}")
        else:
            st.error(f"No subjects found for section '{section}' in Section-Subject-Mapping sheet. " 
                    f"Please ensure the section name is correctly mapped to subjects.")

def get_faculty_workload(username):
    """Calculate faculty workload and organize by months"""
    try:
        # Get faculty sheet data
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        
        # Get faculty name for workload lookup
        user_mask = df_faculty['Username'] == username
        if not user_mask.any():
            return 0, pd.DataFrame()
            
        workload_data = []
        
        # Process each month column
        for col in df_faculty.columns:
            if col not in ['Faculty Name', 'Username', 'Password']:
                # Get entries for this month
                entries = str(df_faculty.loc[user_mask, col].iloc[0]).split('\n') if pd.notna(df_faculty.loc[user_mask, col].iloc[0]) else []
                
                for entry in entries:
                    if pd.notna(entry) and entry.strip():
                        try:
                            parts = entry.strip().split('_')
                            if len(parts) >= 5:  # Ensure we have all required parts
                                date_str = parts[0]
                                time_str = parts[1]
                                period = parts[2]
                                subject = parts[3]
                                section = "_".join(parts[4:])  # Handle sections with underscores in name
                                
                                # Convert date for filtering
                                date_obj = pd.to_datetime(date_str, format='%d/%m/%Y')
                                
                                # Check date range if provided
                                if 'from_date' in st.session_state and 'to_date' in st.session_state:
                                    from_date = pd.to_datetime(st.session_state.from_date)
                                    to_date = pd.to_datetime(st.session_state.to_date)
                                    if not (from_date <= date_obj <= to_date + pd.Timedelta(days=1)):
                                        continue
                                        
                                workload_data.append({
                                    'Date': date_str,
                                    'Time': time_str,
                                    'Period': period,
                                    'Subject': subject,
                                    'Section': section
                                })
                        except Exception as e:
                            st.error(f"Error processing entry: {entry}")
                            continue
        
        if workload_data:
            # Convert to DataFrame and add Month column
            df_workload = pd.DataFrame(workload_data)
            df_workload['DateObj'] = pd.to_datetime(df_workload['Date'], format='%d/%m/%Y')
            df_workload['Month'] = df_workload['DateObj'].dt.strftime('%b%Y')
            
            # Sort by date
            df_workload = df_workload.sort_values('DateObj', ascending=False)
            
            # Remove DateObj column
            df_workload = df_workload.drop('DateObj', axis=1)
            
            return len(df_workload), df_workload
            
        return 0, pd.DataFrame()
        
    except Exception as e:
        st.error(f"Error calculating workload: {str(e)}")
        return 0, pd.DataFrame()

def workload_analysis_page():
    """Page for viewing faculty workload"""
    st.subheader("My Workload Analysis")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.from_date = st.date_input(
            "From Date",
            datetime.now().replace(day=1),
            format="YYYY/MM/DD"
        )
    with col2:
        st.session_state.to_date = st.date_input(
            "To Date",
            datetime.now(),
            format="YYYY/MM/DD"
        )
    
    # Get workload data
    total_periods, workload_df = get_faculty_workload(st.session_state.username)
    
    if not workload_df.empty:
        # Summary metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Classes", total_periods)
        with col2:
            unique_days = workload_df['Date'].nunique()
            st.metric("Days Engaged", unique_days)
        with col3:
            avg_classes = total_periods / max(unique_days, 1)
            st.metric("Daily Average", f"{avg_classes:.1f}")
        
        # Show section and subject breakdown
        st.subheader("Teaching Distribution")
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("##### Subject-wise Classes")
            subject_counts = workload_df['Subject'].value_counts().reset_index()
            subject_counts.columns = ['Subject', 'Classes']
            st.dataframe(subject_counts, hide_index=True)
            
        with col2:
            st.write("##### Section-wise Classes")
            section_counts = workload_df['Section'].value_counts().reset_index()
            section_counts.columns = ['Section', 'Classes']
            st.dataframe(section_counts, hide_index=True)
        
        # Detailed records grouped by month
        st.subheader("Detailed Class Records")
        for month in sorted(workload_df['Month'].unique(), reverse=True):
            with st.expander(f"### {month}"):
                month_data = workload_df[workload_df['Month'] == month].copy()
                month_data = month_data.drop('Month', axis=1)
                st.dataframe(
                    month_data,
                    column_config={
                        'Date': st.column_config.TextColumn('Date', width=100),
                        'Time': st.column_config.TextColumn('Time', width=100),
                        'Period': st.column_config.TextColumn('Period', width=80),
                        'Section': st.column_config.TextColumn('Section', width=150),
                        'Subject': st.column_config.TextColumn('Subject', width=150)
                    },
                    hide_index=True,
                    use_container_width=True
                )
                
                # Add download option for this month's data
                csv = month_data.to_csv(index=False)
                st.download_button(
                    label=f"Download {month} Records",
                    data=csv,
                    file_name=f"workload_{month}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
    else:
        st.info("No classes recorded in the selected date range")


def mark_attendance(section, period, attendance_data, username, subject):
    """Modified mark_attendance function to update faculty log with proper name"""
    try:
        # First get faculty name from username
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        user_row = df_faculty[df_faculty['Username'] == username].iloc[0]
        faculty_name = user_row['Faculty Name']  # This will get the full name with RVIT ID
        
        date_str = datetime.now().strftime('%d/%m/%Y')
        time_str = datetime.now().strftime('%I:%M%p')
        
        unsuccessful_records = []
        
        # Rest of the code remains the same...
        original_sections = {}
        for ht_number, data in attendance_data.items():
            orig_section = data['original_section'].replace("Original: ", "")
            if orig_section not in original_sections:
                original_sections[orig_section] = {}
            original_sections[orig_section][ht_number] = data['status']
        
        success = True
        for orig_section, students in original_sections.items():
            try:
                df = pd.read_excel('attendance.xlsx', sheet_name=orig_section)
                
                with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                    for ht_number, status in students.items():
                        try:
                            row_mask = df['HT Number'] == ht_number
                            if not row_mask.any():
                                student_df = pd.read_excel('attendance.xlsx', sheet_name=section)
                                student_name = student_df[student_df['HT Number'] == ht_number]['Student Name'].iloc[0]
                                unsuccessful_records.append({
                                    'HT Number': ht_number,
                                    'Student Name': student_name,
                                    'Original Section': orig_section,
                                    'Reason': f"Student not found in section {orig_section}"
                                })
                                continue
                            
                            attendance_value = f"{date_str}_{time_str}_{status}_{faculty_name}_{subject}"
                            current_value = df.loc[row_mask, period].iloc[0]
                            df.loc[row_mask, period] = (
                                f"{current_value}\n{attendance_value}" if pd.notna(current_value) and current_value 
                                else attendance_value
                            )
                            
                        except Exception as e:
                            unsuccessful_records.append({
                                'HT Number': ht_number,
                                'Student Name': "Unknown",
                                'Original Section': orig_section,
                                'Reason': "Unable to process attendance"
                            })
                    
                    df.to_excel(writer, sheet_name=orig_section, index=False)
                    
                    # Format worksheet
                    worksheet = writer.sheets[orig_section]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    for column in worksheet.columns:
                        max_length = max(len(str(cell.value or '')) for cell in column)
                        worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                
            except Exception as e:
                success = False
                for ht_number in students.keys():
                    unsuccessful_records.append({
                        'HT Number': ht_number,
                        'Student Name': "Unknown",
                        'Original Section': orig_section,
                        'Reason': "Section access issue"
                    })
        
        # Update faculty log if attendance marking was successful
        if success:
            update_faculty_log(faculty_name, section, period, subject)
        
        return success, unsuccessful_records
    
    except Exception as e:
        st.error(f"Error marking attendance: {str(e)}")
        return False, []


def reset_password():
    """Function to handle password reset with direct field presentation"""
    st.subheader("Reset Password")
    
    username = st.text_input("Username", key="reset_pwd_username")
    current_password = st.text_input("Current Password", type="password", key="reset_pwd_current")
    new_password = st.text_input("New Password", type="password", key="reset_pwd_new")
    confirm_password = st.text_input("Confirm New Password", type="password", key="reset_pwd_confirm")
    
    if st.button("Reset Password", key="reset_pwd_button", type="primary"):
        try:
            if not all([username, current_password, new_password, confirm_password]):
                st.error("All fields are required")
                return
                
            # Read faculty data
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            
            # Verify credentials
            user_mask = (df_faculty['Username'] == username) & \
                       (df_faculty['Password'] == current_password)
            if not user_mask.any():
                st.error("Invalid credentials")
                return
            
            # Verify new passwords match
            if new_password != confirm_password:
                st.error("New passwords do not match")
                return
            
            # Update password
            df_faculty.loc[user_mask, 'Password'] = new_password
            
            # Save changes while preserving all columns
            with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                df_faculty.to_excel(writer, sheet_name='Faculty', index=False)
                
                # Format worksheet
                worksheet = writer.sheets['Faculty']
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Set column widths
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
            
            st.success("Password updated successfully! Please login again.")
            
            # Clear session state to force re-login
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
            
        except Exception as e:
            st.error(f"Error resetting password: {str(e)}")

def reset_username():
    """Function to handle username reset"""
    st.subheader("Reset Username")
    
    current_username = st.text_input("Current Username", key="reset_user_current")
    password = st.text_input("Current Password", type="password", key="reset_user_pwd")
    new_username = st.text_input("New Username", key="reset_user_new")
    
    if st.button("Reset Username", key="reset_user_button", type="primary"):
        try:
            if not all([current_username, password, new_username]):
                st.error("All fields are required")
                return
                
            # Read faculty data
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            
            # Verify credentials
            user_mask = (df_faculty['Username'] == current_username) & \
                       (df_faculty['Password'] == password)
            if not user_mask.any():
                st.error("Invalid credentials")
                return
                
            # Check if new username already exists
            if (df_faculty['Username'] == new_username).any():
                st.error("Username already exists")
                return
                
            # Update username
            df_faculty.loc[user_mask, 'Username'] = new_username
            
            # Save changes while preserving all columns
            with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay') as writer:
                df_faculty.to_excel(writer, sheet_name='Faculty', index=False)
                
                # Format worksheet
                worksheet = writer.sheets['Faculty']
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Set column widths
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
            
            st.success("Username updated successfully! Please login again with your new username.")
            
            # Clear session state to force re-login
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
            
        except Exception as e:
            st.error(f"Error resetting username: {str(e)}")

def main():
    if 'logged_in' not in st.session_state:
        st.title("Login")
        
        # Add tabs for login and reset options
        tab1, tab2, tab3 = st.tabs(["Login", "Reset Password", "Reset Username"])
        
        with tab1:
            login_type = st.radio("Select Login Type", ["Faculty", "Admin"], key="login_type")
            username = st.text_input("Username", key="login_username")
            password = st.text_input("Password", type="password", key="login_password")
            
            if st.button("Login", key="login_button", type="primary"):
                if check_login(username, password, login_type == "Admin"):
                    st.session_state.logged_in = True
                    st.session_state.is_admin = (login_type == "Admin")
                    st.session_state.username = username
                    
                    # Get faculty name for display
                    df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
                    user_row = df_faculty[df_faculty['Username'] == username].iloc[0]
                    st.session_state.faculty_name = user_row['Faculty Name']
                    
                    st.rerun()
                else:
                    st.error("Invalid credentials")
        
        with tab2:
            reset_password()
            
        with tab3:
            reset_username()
    else:
        if st.session_state.is_admin:
            admin_page()
        else:
            faculty_page()


if __name__ == "__main__":
    main()