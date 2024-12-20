import streamlit as st
import pandas as pd
from datetime import datetime
import hashlib
import openpyxl
from openpyxl.styles import Alignment
import io
import numpy as np

# Configure Streamlit page
st.set_page_config(page_title="Attendance Management System", layout="wide")

# Admin credentials
ADMIN_CREDENTIALS = {
    "a": hashlib.sha256("a".encode()).hexdigest()
}

def check_login(username, password, is_admin=False):
    """Verify login credentials"""
    try:
        if is_admin:
            hashed_password = hashlib.sha256(password.encode()).hexdigest()
            return ADMIN_CREDENTIALS.get(username) == hashed_password
        else:
            df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
            return any((df_faculty['Faculty Name'] == username) & 
                      (df_faculty['Password'] == password))
    except Exception as e:
        st.error(f"Login error: {str(e)}")
        return False

def get_section_subjects(section):
    """Get subjects mapped to a specific section"""
    try:
        df_mapping = pd.read_excel('attendance.xlsx', sheet_name='Section-Subject-Mapping')
        section_row = df_mapping[df_mapping['Section'] == section].iloc[0]
        subjects = [s.strip() for s in str(section_row['Subject Names']).split('\n') if s.strip()]
        return subjects
    except Exception as e:
        st.error(f"Error getting subjects: {str(e)}")
        return []

def get_sections(for_attendance=False):
    """
    Get sections based on context
    for_attendance=True: Returns only manipulated sections (without O prefix)
    for_attendance=False: Returns original sections (with O prefix)
    """
    try:
        all_sheets = pd.ExcelFile('attendance.xlsx').sheet_names
        if for_attendance:
            # For attendance marking: Only show manipulated sections
            return [s for s in all_sheets if not s.startswith('(O)') 
                   and s not in ['Faculty', 'Section-Subject-Mapping']]
        else:
            # For other features: Show original sections
            return [s.replace('(O)', '') for s in all_sheets 
                   if s.startswith('(O)')]
    except Exception as e:
        st.error(f"Error getting sections: {str(e)}")
        return []


def mark_attendance_page():
    """Page for marking attendance"""
    section = st.session_state.sections[0] if st.session_state.sections else None
    subject = st.session_state.subject
    period = st.session_state.period

    if section and subject and period:
        df_students = get_student_data(section)
        if df_students is not None:
            attendance_data = {}
            
            # Select all/none buttons
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Select All"):
                    st.session_state.select_all = True
            with col2:
                if st.button("Select None"):
                    st.session_state.select_all = False
            
            st.write("### Students")
            for _, student in df_students.iterrows():
                col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                with col1:
                    st.write(student['HT Number'])
                with col2:
                    st.write(student['Student Name'])
                with col3:
                    st.write(student['Original Section'])
                with col4:
                    default_value = getattr(st.session_state, 'select_all', True)
                    status = 'P' if st.checkbox("Present", key=student['HT Number'], value=default_value) else 'A'
                    attendance_data[student['HT Number']] = {
                        'status': status,
                        'original_section': student['Original Section']
                    }
            
            if st.button("Submit Attendance"):
                success, unsuccessful_records = mark_attendance(section, period, attendance_data, st.session_state.username, subject)
                
                if unsuccessful_records:
                    # Show notice for unsuccessful records
                    st.write("### Attendance Status")
                    st.info(f"Successfully recorded attendance for {len(attendance_data) - len(unsuccessful_records)} students")
                    
                    st.warning("Notice: Following students have attendance marking issues:")
                    unsuccessful_df = pd.DataFrame(unsuccessful_records)
                    st.dataframe(
                        unsuccessful_df,
                        column_config={
                            'HT Number': st.column_config.TextColumn('HT Number', width=120),
                            'Student Name': st.column_config.TextColumn('Student Name', width=180),
                            'Original Section': st.column_config.TextColumn('Original Section', width=150),
                            'Reason': st.column_config.TextColumn('Issue Details', width=300)
                        },
                        hide_index=True
                    )
                else:
                    st.success(f"Successfully recorded attendance for all {len(attendance_data)} students")


# Updated mark_attendance function to return both success status and unsuccessful records
def update_faculty_log(faculty_name, section, period, subject):
    """Update faculty attendance log with month-wise organization"""
    try:
        # Read faculty sheet
        df = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        
        # Get current month-year
        current_date = datetime.now()
        month_year = current_date.strftime('%b%Y')
        
        # Create log entry in the format: DD/MM/YYYY_HH:MM[AM/PM]_[Period]_[Subject]_[Section]
        log_entry = f"{current_date.strftime('%d/%m/%Y_%I:%M%p')}_{period}_{subject}_{section}"
        
        # Check if month-year column exists, if not create it
        if month_year not in df.columns:
            # Get existing columns
            existing_cols = list(df.columns)
            # Find the position after 'Password' column
            password_idx = existing_cols.index('Password')
            # Insert new column after Password
            df.insert(password_idx + 1, month_year, '')
        
        # Update the log for the faculty
        faculty_mask = df['Faculty Name'] == faculty_name
        if faculty_mask.any():
            current_log = df.loc[faculty_mask, month_year].iloc[0]
            # Add new entry with proper newline handling
            if pd.notna(current_log) and str(current_log).strip():
                new_log = f"{current_log}\n{log_entry}"
            else:
                new_log = log_entry
            df.loc[faculty_mask, month_year] = new_log
        
        # Save the updated sheet while preserving column order
        with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Faculty', index=False)
            
            # Format the worksheet
            worksheet = writer.sheets['Faculty']
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Set column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                adjusted_width = min(50, max(12, max_length + 2))
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        return True
    except Exception as e:
        st.error(f"Error updating faculty log: {str(e)}")
        return False

def get_faculty_workload(username):
    """Calculate faculty workload and organize by months"""
    try:
        # Get faculty sheet data
        df_faculty = pd.read_excel('attendance.xlsx', sheet_name='Faculty')
        faculty_row = df_faculty[df_faculty['Faculty Name'] == username]
        
        if not faculty_row.empty:
            workload_data = []
            
            # Process each month column
            for col in faculty_row.columns:
                if col not in ['Faculty Name', 'Password']:
                    entries = str(faculty_row[col].iloc[0]).split('\n') if pd.notna(faculty_row[col].iloc[0]) else []
                    
                    for entry in entries:
                        if entry.strip():  # Only process non-empty entries
                            try:
                                # Parse the entry: DD/MM/YYYY_HH:MM[AM/PM]_P2_BCME_B.Tech-I-CSE-A
                                parts = entry.strip().split('_')
                                if len(parts) >= 5:
                                    date_str = parts[0]
                                    time_str = parts[1]
                                    period = parts[2]
                                    subject = parts[3]
                                    section = parts[4]
                                    
                                    workload_data.append({
                                        'Date': date_str,
                                        'Time': time_str,
                                        'Period': period,
                                        'Section': section,
                                        'Subject': subject
                                    })
                            except Exception as e:
                                st.error(f"Error processing entry: {entry}, Error: {str(e)}")
                                continue
            
            # Convert to DataFrame
            df_workload = pd.DataFrame(workload_data)
            
            if not df_workload.empty:
                # Convert Date column to datetime for filtering
                df_workload['DateObj'] = pd.to_datetime(df_workload['Date'], format='%d/%m/%Y')
                
                # Filter based on date range if provided
                if 'from_date' in st.session_state and 'to_date' in st.session_state:
                    from_date = pd.to_datetime(st.session_state.from_date)
                    to_date = pd.to_datetime(st.session_state.to_date)
                    
                    # Filter using DateObj column
                    df_workload = df_workload[
                        (df_workload['DateObj'] >= from_date) & 
                        (df_workload['DateObj'] <= to_date)
                    ]
                
                if not df_workload.empty:
                    # Add Month column for grouping
                    df_workload['Month'] = df_workload['DateObj'].dt.strftime('%b%Y')
                    # Sort by date
                    df_workload = df_workload.sort_values('DateObj')
                    # Drop the DateObj column as it's no longer needed
                    df_workload = df_workload.drop('DateObj', axis=1)
                    
                    return len(df_workload), df_workload
            
        return 0, pd.DataFrame()
    
    except Exception as e:
        st.error(f"Error calculating workload: {str(e)}")
        return 0, pd.DataFrame()

def workload_analysis_page():
    """Page for viewing faculty workload"""
    st.subheader("My Workload Analysis")
    
    # Date range selection
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.from_date = st.date_input(
            "From Date",
            datetime.now().replace(day=1),
            format="YYYY/MM/DD"
        )
    with col2:
        st.session_state.to_date = st.date_input(
            "To Date",
            datetime.now(),
            format="YYYY/MM/DD"
        )
    
    # Get workload data
    total_periods, workload_df = get_faculty_workload(st.session_state.username)
    
    if not workload_df.empty:
        # Summary metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Classes", total_periods)
        with col2:
            unique_days = workload_df['Date'].nunique()
            st.metric("Days Engaged", unique_days)
        with col3:
            avg_classes = total_periods / max(unique_days, 1)
            st.metric("Daily Average", f"{avg_classes:.1f}")
        
        # Show section and subject breakdown
        st.subheader("Teaching Distribution")
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("##### Subject-wise Classes")
            subject_counts = workload_df['Subject'].value_counts().reset_index()
            subject_counts.columns = ['Subject', 'Classes']
            st.dataframe(subject_counts, hide_index=True)
            
        with col2:
            st.write("##### Section-wise Classes")
            section_counts = workload_df['Section'].value_counts().reset_index()
            section_counts.columns = ['Section', 'Classes']
            st.dataframe(section_counts, hide_index=True)
        
        # Detailed records grouped by month
        st.subheader("Detailed Class Records")
        for month in workload_df['Month'].unique():
            with st.expander(f"### {month}"):
                month_data = workload_df[workload_df['Month'] == month].copy()
                month_data = month_data.drop('Month', axis=1)
                st.dataframe(
                    month_data.sort_values('Date'),
                    column_config={
                        'Date': st.column_config.TextColumn('Date', width=100),
                        'Time': st.column_config.TextColumn('Time', width=100),
                        'Period': st.column_config.TextColumn('Period', width=80),
                        'Section': st.column_config.TextColumn('Section', width=150),
                        'Subject': st.column_config.TextColumn('Subject', width=150)
                    },
                    hide_index=True,
                    use_container_width=True
                )
    else:
        st.info("No classes recorded in the selected date range")  
  
def create_template_df(sheet_name):
    """Create template DataFrame based on sheet type with updated Faculty template"""
    if sheet_name == 'Faculty':
        # Get current month-year
        current_month = datetime.now().strftime('%b%Y')
        return pd.DataFrame(columns=['Faculty Name', 'Password', current_month])
    elif sheet_name == 'Section-Subject-Mapping':
        return pd.DataFrame(columns=['Section', 'Subject Names'])
    elif sheet_name.startswith('(O)'):
        return pd.DataFrame(columns=['HT Number', 'Student Name', 'P1', 'P2', 'P3', 'P4', 'P5', 'P6'])
    else:  # Manipulated sections
        return pd.DataFrame(columns=['HT Number', 'Student Name', 'Original Section'])

def mark_attendance(section, period, attendance_data, faculty_name, subject):
    """Modified mark_attendance function to update faculty log"""
    try:
        date_str = datetime.now().strftime('%d/%m/%Y')
        time_str = datetime.now().strftime('%I:%M%p')
        
        unsuccessful_records = []
        
        # Group students by their original sections
        original_sections = {}
        for ht_number, data in attendance_data.items():
            orig_section = data['original_section'].replace("Original: ", "")
            if orig_section not in original_sections:
                original_sections[orig_section] = {}
            original_sections[orig_section][ht_number] = data['status']
        
        # Mark attendance in each original section
        success = True
        for orig_section, students in original_sections.items():
            try:
                df = pd.read_excel('attendance.xlsx', sheet_name=orig_section)
                
                with pd.ExcelWriter('attendance.xlsx', mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                    for ht_number, status in students.items():
                        try:
                            row_mask = df['HT Number'] == ht_number
                            if not row_mask.any():
                                student_df = pd.read_excel('attendance.xlsx', sheet_name=section)
                                student_name = student_df[student_df['HT Number'] == ht_number]['Student Name'].iloc[0]
                                unsuccessful_records.append({
                                    'HT Number': ht_number,
                                    'Student Name': student_name,
                                    'Original Section': orig_section,
                                    'Reason': f"Student not found in section {orig_section}"
                                })
                                continue
                            
                            attendance_value = f"{date_str}_{time_str}_{status}_{faculty_name}_{subject}"
                            current_value = df.loc[row_mask, period].iloc[0]
                            df.loc[row_mask, period] = (
                                f"{current_value}\n{attendance_value}" if pd.notna(current_value) and current_value 
                                else attendance_value
                            )
                            
                        except Exception as e:
                            unsuccessful_records.append({
                                'HT Number': ht_number,
                                'Student Name': "Unknown",
                                'Original Section': orig_section,
                                'Reason': "Unable to process attendance"
                            })
                    
                    df.to_excel(writer, sheet_name=orig_section, index=False)
                    
                    # Format worksheet
                    worksheet = writer.sheets[orig_section]
                    for row in worksheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    for column in worksheet.columns:
                        max_length = max(len(str(cell.value or '')) for cell in column)
                        worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                
            except Exception as e:
                success = False
                for ht_number in students.keys():
                    unsuccessful_records.append({
                        'HT Number': ht_number,
                        'Student Name': "Unknown",
                        'Original Section': orig_section,
                        'Reason': "Section access issue"
                    })
        
        # Update faculty log if attendance marking was successful
        if success:
            update_faculty_log(faculty_name, section, period, subject)
        
        return success, unsuccessful_records
    
    except Exception as e:
        return False, []

def get_student_data(section):
    """Get student data for a manipulated section"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=section)
        students_df = df[['HT Number', 'Student Name', 'Original Section']].fillna('')
        # Format original section display without duplication
        students_df['Original Section'] = students_df['Original Section'].apply(
            lambda x: f"Original: {x}" if not x.startswith("Original:") else x
        )
        return students_df
    except Exception as e:
        st.error(f"Error getting student data: {str(e)}")
        return None

def faculty_page():
    st.title(f"Welcome, {st.session_state.username}")
    
    with st.sidebar:
        st.header("Navigation")
        page = st.radio(
            "Select", 
            ["Mark Attendance", "View Statistics", "Student Reports", 
             "Subject Analysis", "My Workload"]
        )

        if page == "Mark Attendance":
            # Period selection
            st.session_state.period = st.selectbox(
                "Select Period",
                options=[''] + ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']
            )
            
            # Section selection - only manipulated sections for attendance
            sections = get_sections(for_attendance=True)
            st.session_state.sections = st.multiselect(
                "Select Section(s)",
                options=sections,
                max_selections=2,
                help="Select up to 2 sections"
            )
            
            # Show available subjects for selected section
            if st.session_state.sections:
                all_subjects = []
                for section in st.session_state.sections:
                    subjects = get_section_subjects(section)
                    all_subjects.extend(subjects)
                # Remove duplicates and sort
                unique_subjects = sorted(list(set(all_subjects)))
                
                st.session_state.subject = st.selectbox(
                    "Select Subject",
                    options=unique_subjects if unique_subjects else [''],
                    help="Select the subject being taught"
                )
        
        if st.button("Logout"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()

    if page == "Mark Attendance":
        if not (st.session_state.get('period') and st.session_state.get('sections') 
                and st.session_state.get('subject')):
            st.info("Please select Period, Section(s), and Subject from the sidebar")
            return
            
        mark_attendance_page()
    elif page == "View Statistics":
        view_statistics_page()
    elif page == "Student Reports":
        student_reports_page()
    elif page == "Subject Analysis":
        subject_analysis_page()
    else:  # Workload Analysis
        workload_analysis_page()





def subject_analysis_page():
    """Page for subject-wise analysis"""
    st.subheader("Subject-wise Analysis")
    
    # Use original sections for analysis
    sections = get_sections(for_attendance=False)
    section = st.selectbox("Select Section", sections)
    
    if section:
        subjects = get_section_subjects(section)
        subject = st.selectbox("Select Subject", subjects)
        
        if subject:
            analysis_df = get_subject_analysis(section, subject)
            if not analysis_df.empty:
                st.write("### Subject Statistics")
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    avg_attendance = analysis_df['Attendance %'].mean()
                    st.metric("Average Attendance", f"{avg_attendance:.2f}%")
                with col2:
                    total_classes = analysis_df['Total Classes'].max()
                    st.metric("Total Classes", total_classes)
                with col3:
                    below_75 = len(analysis_df[analysis_df['Attendance %'] < 75])
                    st.metric("Students Below 75%", below_75)
                
                st.write("### Student-wise Analysis")
                st.dataframe(analysis_df)
                
                if st.button("Download Analysis"):
                    csv = analysis_df.to_csv(index=False)
                    st.download_button(
                        label="Download CSV",
                        data=csv,
                        file_name=f"subject_analysis_{section}_{subject}_{datetime.now().strftime('%Y%m%d')}.csv",
                        mime="text/csv"
                    )
            else:
                st.info("No data available for analysis")

def student_reports_page():
    """Page for individual student reports"""
    st.subheader("Individual Student Reports")
    
    # Use original sections for reports
    sections = get_sections(for_attendance=False)
    section = st.selectbox("Select Section", sections)
    
    if section:
        df_students = get_student_data(section)
        if df_students is not None:
            student = st.selectbox(
                "Select Student", 
                df_students['HT Number'].tolist(),
                format_func=lambda x: f"{x} - {df_students[df_students['HT Number']==x]['Student Name'].iloc[0]}"
            )
            
            if student:
                student_data = df_students[df_students['HT Number'] == student].iloc[0]
                
                st.write(f"### Attendance Report for {student}")
                st.write(f"**Name:** {student_data['Student Name']}")
                st.write(f"**Original Section:** {student_data['Original Section']}")
                
                # Get attendance details
                attendance_data = get_student_attendance_details(student_data['Original Section'].replace("Original: ", ""), student)
                
                # Properly check if attendance data exists and is not empty
                if attendance_data is not None and not attendance_data.empty:
                    st.dataframe(attendance_data.sort_values('Date', ascending=False))
                    
                    if st.button("Download Student Report"):
                        csv = attendance_data.to_csv(index=False)
                        st.download_button(
                            label="Download CSV",
                            data=csv,
                            file_name=f"student_report_{student}_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                else:
                    st.info("No attendance records found")

def get_student_attendance_details(section, student_id):
    """Get detailed attendance records for a student"""
    try:
        # Remove (O) prefix if present
        section = section.replace("(O)", "").strip()
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        student_row = df[df['HT Number'] == student_id]
        
        if student_row.empty:
            return None
            
        student_row = student_row.iloc[0]
        attendance_data = []
        
        for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
            if pd.notna(student_row[period]) and student_row[period]:
                entries = str(student_row[period]).split('\n')
                for entry in entries:
                    if entry.strip():
                        try:
                            date, time, status, faculty, subject = entry.split('_')
                            attendance_data.append({
                                'Date': date,
                                'Time': time,
                                'Period': period,
                                'Status': status,
                                'Faculty': faculty,
                                'Subject': subject
                            })
                        except ValueError:
                            continue  # Skip malformed entries
        
        if not attendance_data:  # If no attendance records found
            return pd.DataFrame()  # Return empty DataFrame
            
        return pd.DataFrame(attendance_data)
    except Exception as e:
        st.error(f"Error getting attendance details: {str(e)}")
        return None


def get_subject_analysis(section, subject):
    """Get subject-wise attendance analysis"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        analysis = []
        
        for _, row in df.iterrows():
            present = 0
            total = 0
            
            for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                if pd.notna(row[period]) and row[period]:
                    entries = str(row[period]).split('\n')
                    for entry in entries:
                        if subject in entry:
                            total += 1
                            if '_P_' in entry:
                                present += 1
            
            if total > 0:
                percentage = (present / total) * 100
            else:
                percentage = 0
                
            analysis.append({
                'HT Number': row['HT Number'],
                'Student Name': row['Student Name'],
                'Classes Attended': present,
                'Total Classes': total,
                'Attendance %': round(percentage, 2)
            })
        
        return pd.DataFrame(analysis)
    except Exception as e:
        st.error(f"Error in subject analysis: {str(e)}")
        return pd.DataFrame()




def check_duplicate_attendance(section, period, date_str):
    """Check if attendance is already marked for given section, period and date"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        if period in df.columns:
            for value in df[period].dropna():
                for entry in str(value).split('\n'):
                    if entry.startswith(date_str):
                        return True
        return False
    except Exception as e:
        st.error(f"Error checking duplicate attendance: {str(e)}")
        return True



def get_attendance_stats(section, from_date=None, to_date=None):
    """Calculate attendance statistics with attended/conducted format"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=f'(O){section}')
        subjects = get_section_subjects(section)
        
        stats = []
        for _, row in df.iterrows():
            student_stats = {
                'HT Number': row['HT Number'],
                'Student Name': row['Student Name']
            }
            
            total_present = 0
            total_classes = 0
            
            # Calculate for each mapped subject
            for subject in subjects:
                present = 0
                total = 0
                
                for period in ['P1', 'P2', 'P3', 'P4', 'P5', 'P6']:
                    if pd.notna(row[period]) and row[period]:
                        entries = str(row[period]).split('\n')
                        for entry in entries:
                            if subject in entry:
                                total += 1
                                total_classes += 1
                                if '_P_' in entry:
                                    present += 1
                                    total_present += 1
                
                # Only add subject column if there are classes for this subject
                if total > 0:
                    student_stats[f"{subject}\n(Attended/Conducted)"] = f"{present}/{total}"
            
            # Add total attended/conducted before Overall %
            student_stats[f"Total\n(Attended/Conducted)"] = f"{total_present}/{total_classes}"
            
            # Calculate overall percentage
            overall_percentage = (total_present / total_classes * 100) if total_classes > 0 else 0
            student_stats['Overall %'] = round(overall_percentage, 2)
            
            stats.append(student_stats)
        
        # Convert to DataFrame
        stats_df = pd.DataFrame(stats)
        
        # Ensure correct column order
        base_columns = ['HT Number', 'Student Name']
        subject_columns = [col for col in stats_df.columns 
                         if '(Attended/Conducted)' in col and 'Total' not in col]
        total_column = [col for col in stats_df.columns if col.startswith('Total')]
        percentage_column = ['Overall %']
        
        # Reorder columns
        ordered_columns = base_columns + subject_columns + total_column + percentage_column
        stats_df = stats_df[ordered_columns]
        
        return stats_df
    except Exception as e:
        st.error(f"Error calculating statistics: {str(e)}")
        return None

def view_statistics_page():
    """Page for viewing attendance statistics"""
    st.subheader("View Attendance Statistics")
    
    # Get original sections only
    sections = get_sections(for_attendance=False)
    section = st.selectbox("Select Section", sections)
    
    col1, col2 = st.columns(2)
    with col1:
        from_date = st.date_input("From Date")
    with col2:
        to_date = st.date_input("To Date")
    
    if section:
        stats_df = get_attendance_stats(
            section, 
            from_date.strftime('%Y-%m-%d') if from_date else None,
            to_date.strftime('%Y-%m-%d') if to_date else None
        )
        
        if stats_df is not None and not stats_df.empty:
            st.write("### Overall Statistics")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Students", len(stats_df))
            with col2:
                avg_attendance = stats_df['Overall %'].mean()
                st.metric("Average Attendance", f"{avg_attendance:.2f}%")
            with col3:
                below_75 = len(stats_df[stats_df['Overall %'] < 75])
                st.metric("Students Below 75%", below_75)
            
            # Configure column display
            column_config = {
                'HT Number': st.column_config.TextColumn('HT Number', width=120),
                'Student Name': st.column_config.TextColumn('Student Name', width=180),
                'Overall %': st.column_config.NumberColumn(
                    'Overall %',
                    format="%.2f%%",
                    width=100
                )
            }
            
            # Add configuration for subject and total columns
            for col in stats_df.columns:
                if 'Attended/Conducted' in col:
                    column_config[col] = st.column_config.TextColumn(
                        col,
                        width=150
                    )
            
            st.write("### Student-wise Statistics")
            st.dataframe(
                stats_df,
                column_config=column_config,
                use_container_width=True,
                hide_index=True
            )
            
            if st.button("Download Report"):
                csv = stats_df.to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name=f"attendance_stats_{section}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )
        else:
            st.info("No attendance records found for the selected criteria")

def get_column_width(col_name, values):
    try:
        max_length = max(
            max(len(str(val)) for val in values if val is not None),
            len(str(col_name))
        )
        return min(max_length * 10, 300)
    except:
        return 150



def validate_upload_data(df, sheet_name):
    """Validate uploaded data against expected format"""
    template_df = create_template_df(sheet_name)
    return all(col in df.columns for col in template_df.columns)

def show_data_editor(sheet):
    """Show the data editor component with improved layout"""
    try:
        if sheet == 'Faculty':
            # Special handling for Faculty sheet
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet, usecols=['Faculty Name', 'Password'])
            df = df.fillna('')
        else:
            # Regular handling for other sheets
            df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
            df = df.convert_dtypes().fillna('')
        
        # Get actual number of rows
        data_rows = len(df)
        
        # Configure columns
        column_config = {}
        for col in df.columns:
            width = 150 if col in ['HT Number', 'Student Name', 'Faculty Name'] else 300
            column_config[str(col)] = st.column_config.TextColumn(
                col,
                width=width,
                help=f"Enter {col}",
                max_chars=None
            )
        
        # Display editor with minimal extra rows
        edited_df = st.data_editor(
            df,
            use_container_width=True,
            num_rows="dynamic",
            column_config=column_config,
            hide_index=True,
            height=min(600, (data_rows + 2) * 35)
        )
        
        # Create columns for buttons
        col1, col2, col3 = st.columns([1, 1, 2])
        
        with col1:
            if st.button("Save Changes"):
                try:
                    with pd.ExcelWriter('attendance.xlsx', 
                                      mode='a', 
                                      if_sheet_exists='overlay',
                                      engine='openpyxl') as writer:
                        edited_df.to_excel(writer, sheet_name=sheet, index=False)
                        
                        # Format worksheet
                        worksheet = writer.sheets[sheet]
                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        # Set column widths
                        for column in worksheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            adjusted_width = min(50, max(12, max_length + 2))
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    st.success("Changes saved successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error saving changes: {str(e)}")

        with col2:
            # Download current sheet button
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                edited_df.to_excel(writer, index=False)
                worksheet = writer.sheets['Sheet1']
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value or '')) for cell in column)
                    worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
            
            st.download_button(
                label="Download Sheet",
                data=buffer.getvalue(),
                file_name=f"{sheet}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
                
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")

def get_faculty_data(sheet):
    """Get faculty data with proper type conversion"""
    try:
        df = pd.read_excel('attendance.xlsx', sheet_name=sheet)
        # Convert all columns to string type and handle dates
        for col in df.columns:
            df[col] = df[col].apply(lambda x: str(x) if pd.notnull(x) else '')
        return df
    except Exception as e:
        st.error(f"Error getting faculty data: {str(e)}")
        return None


def admin_page():
    """Admin dashboard page with improved error handling"""
    st.title("Admin Dashboard")
    
    try:
        with st.sidebar:
            st.header("Data Management")
            sheet = st.selectbox("Select Sheet", pd.ExcelFile('attendance.xlsx').sheet_names)
            
            st.write("### Download Options")
            
            # Upload complete workbook button
            uploaded_workbook = st.file_uploader("Upload Complete Workbook", type=['xlsx'])
            if uploaded_workbook is not None:
                try:
                    xls = pd.ExcelFile(uploaded_workbook)
                    st.write("Sheets found in workbook:", xls.sheet_names)
                    
                    if st.button("Confirm Upload"):
                        with pd.ExcelWriter('attendance.xlsx', 
                                          mode='a', 
                                          if_sheet_exists='overlay') as writer:
                            for sheet_name in xls.sheet_names:
                                df = pd.read_excel(uploaded_workbook, sheet_name=sheet_name)
                                df.to_excel(writer, sheet_name=sheet_name, index=False)
                        st.success("Workbook uploaded successfully!")
                        st.rerun()
                except Exception as e:
                    st.error(f"Error uploading workbook: {str(e)}")
            
            # Download complete workbook
            if st.button("Download Complete Workbook"):
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    xls = pd.ExcelFile('attendance.xlsx')
                    for sheet_name in xls.sheet_names:
                        df = pd.read_excel('attendance.xlsx', sheet_name=sheet_name, dtype=str)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        worksheet = writer.sheets[sheet_name]
                        for column in worksheet.columns:
                            max_length = max(len(str(cell.value or '')) for cell in column)
                            worksheet.column_dimensions[column[0].column_letter].width = min(50, max(12, max_length + 2))
                
                st.download_button(
                    label="📥 Download Complete Workbook",
                    data=buffer.getvalue(),
                    file_name="attendance_workbook.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            if st.button("Logout"):
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()
        
        # Add tabs for different operations
        tab1, tab2 = st.tabs(["Edit Data", "Bulk Upload"])
        
        with tab1:
            st.subheader(f"Edit {sheet}")
            show_data_editor(sheet)
        
        with tab2:
            st.subheader("Bulk Upload")
            show_bulk_upload(sheet)
            
    except Exception as e:
        st.error(f"An error occurred: {str(e)}")


def show_bulk_upload(sheet):
    """Show bulk upload interface"""
    # Download template button
    if st.button("Download Template"):
        template_df = create_template_df(sheet)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            template_df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
        st.download_button(
            label="📥 Download Template Excel",
            data=buffer.getvalue(),
            file_name=f"{sheet}_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Upload file
    uploaded_file = st.file_uploader("Upload Excel file", type=['xlsx'])
    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            st.write("Preview of uploaded data:")
            st.dataframe(df)
            
            if st.button("Confirm Upload"):
                if validate_upload_data(df, sheet):
                    with pd.ExcelWriter('attendance.xlsx', 
                                      mode='a',
                                      if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, sheet_name=sheet, index=False)
                    st.success("Data uploaded successfully!")
                    st.rerun()
                else:
                    st.error("Invalid data format. Please use the template.")
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")


def main():
    if 'logged_in' not in st.session_state:
        st.title("Login")
        
        login_type = st.radio("Select Login Type", ["Faculty", "Admin"])
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        
        if st.button("Login"):
            if check_login(username, password, login_type == "Admin"):
                st.session_state.logged_in = True
                st.session_state.is_admin = (login_type == "Admin")
                st.session_state.username = username
                st.rerun()
            else:
                st.error("Invalid credentials")
    else:
        if st.session_state.is_admin:
            admin_page()
        else:
            faculty_page()

if __name__ == "__main__":
    main()